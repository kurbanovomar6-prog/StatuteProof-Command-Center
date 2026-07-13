"""Tests for app/change_register.py — Regulatory Change Register export.

Covers: the three-way join (canonical evidence + human assessment + action-log
decision), CSV/XLSX/HTML rendering, mandatory disclaimer, forbidden-claims
guard, and the empty-but-valid register for a range that matches nothing.
"""

from __future__ import annotations

import hashlib
import io
import json
from pathlib import Path

import pytest

from app.change_register import (
    ChangeRegisterError,
    assert_no_forbidden_claims,
    build_change_register_export,
    build_change_register_rows,
    render_change_register_csv,
    render_change_register_html,
    render_change_register_xlsx,
    validate_register_date_range,
)
from app.evidence_assessment import LEGAL_DISCLAIMER


# ── fixtures ───────────────────────────────────────────────────────────────────

def _write(path: Path, text: str) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _make_canonical_record(
    base: Path,
    *,
    source_id: str = "AE-dfsa-test",
    regulator_slug: str = "dfsa",
    regulator: str = "DFSA",
    source_name: str = "DFSA Test Source",
    run_id: str = "run_AE-dfsa-test_20260620T000000Z",
    record_id: str = "evr_dfsa_test_20260620T000000Z",
    timestamp: str = "2026-06-20T00:00:00Z",
    summary: str = "Saved official-source snapshot changed against the previous normalized content.",
) -> dict:
    """Write a complete canonical evidence record on disk and return it."""
    record_dir = base / "evidence" / regulator_slug / source_id / run_id
    current_hash = _write(record_dir / "current.normalized.txt", f"Current official {regulator} text")
    previous_hash = _write(record_dir / "previous.normalized.txt", f"Previous official {regulator} text")
    _write(record_dir / "raw.html", f"<main>Current official {regulator} text</main>")
    _write(record_dir / "snapshot.html", f"<html><body><main>Current official {regulator} text</main></body></html>")
    _write(record_dir / "metadata.json", json.dumps({"provider": "fixture"}, sort_keys=True))
    _write(
        record_dir / "diff.txt",
        f"- Previous official {regulator} text\n+ Current official {regulator} text\n",
    )

    record = {
        "schema_version": "2.0",
        "record_id": record_id,
        "record_status": "complete",
        "source": {
            "source_id": source_id,
            "regulator": regulator,
            "official_url": f"https://www.{regulator_slug}.ae/example",
            "source_name": source_name,
        },
        "run": {"run_id": run_id, "timestamp": timestamp, "status": "CHANGED"},
        "content": {
            "previous_hash": f"sha256:{previous_hash}",
            "current_hash": f"sha256:{current_hash}",
            "raw_content_path": str((record_dir / "raw.html").relative_to(base)),
            "normalized_current_path": str((record_dir / "current.normalized.txt").relative_to(base)),
            "normalized_previous_path": str((record_dir / "previous.normalized.txt").relative_to(base)),
        },
        "change": {
            "diff_path": str((record_dir / "diff.txt").relative_to(base)),
            "summary": summary,
            "lines_added": 1,
            "lines_removed": 1,
        },
        "files": {
            "snapshot_path": str((record_dir / "snapshot.html").relative_to(base)),
            "raw_path": str((record_dir / "raw.html").relative_to(base)),
            "normalized_path": str((record_dir / "current.normalized.txt").relative_to(base)),
            "previous_path": str((record_dir / "previous.normalized.txt").relative_to(base)),
            "metadata_path": str((record_dir / "metadata.json").relative_to(base)),
            "diff_path": str((record_dir / "diff.txt").relative_to(base)),
        },
        "integrity": {
            "hash_verified": True,
            "integrity_status": "VERIFIED",
            "verified_at": "2026-06-20T00:01:00Z",
        },
        "review": {
            "human_review_required": True,
            "review_status": "approved",
            "review_reason": "Customer-facing brief requires human review.",
        },
    }
    (record_dir / "evidence-record.json").write_text(
        json.dumps(record, indent=2, sort_keys=True), encoding="utf-8"
    )
    return record


def _write_assessment(
    base: Path,
    *,
    evidence_record_id: str,
    impact_level: str = "policy_review",
    reviewer_name: str = "Jane Compliance",
    next_action: str = "Route to MLRO for internal review",
    assessment_status: str = "acknowledged",
) -> dict:
    """Append an assessment row directly (join is by evidence_record_id/run_id)."""
    row = {
        "assessment_id": f"assessment-{evidence_record_id[:12]}",
        "evidence_record_id": evidence_record_id,
        "impact_level": impact_level,
        "reviewer_name": reviewer_name,
        "next_action": next_action,
        "assessment_status": assessment_status,
        "reviewed_at": "2026-06-20T01:00:00Z",
        "legal_disclaimer": LEGAL_DISCLAIMER,
    }
    path = base / "data" / "evidence_assessments" / "assessments.jsonl"
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(row, sort_keys=True) + "\n")
    return row


def _action_lookup(mapping: dict[str, list[dict]]):
    """Return an injectable action_log_lookup over an alert_id → entries map."""
    def _lookup(user_id: int, alert_id: str) -> list[dict]:
        return mapping.get(alert_id, [])
    return _lookup


# ── validate_register_date_range ───────────────────────────────────────────────

def test_validate_register_date_range_allows_empty_bounds():
    ok, msg = validate_register_date_range("", "")
    assert ok is True
    assert msg == ""


def test_validate_register_date_range_rejects_from_after_to():
    ok, msg = validate_register_date_range("2026-06-30", "2026-01-01")
    assert ok is False
    assert msg


def test_validate_register_date_range_rejects_bad_format():
    ok, msg = validate_register_date_range("30/06/2026", "")
    assert ok is False
    assert msg


# ── the join ────────────────────────────────────────────────────────────────────

def _real_draft_alert_id(record: dict) -> str:
    """Re-derive the dashboard's real alert id for a canonical record fixture.

    Mirrors app.alert_drafts.make_draft_alert_id, seeded from the BARE normalized
    hash (the record stores it as ``sha256:<hex>``; the draft id used the bare
    hex). This is the id a customer's act/monitor/no_action decision is keyed by.
    """
    from app.alert_drafts import make_draft_alert_id

    bare_hash = record["content"]["current_hash"].removeprefix("sha256:")
    return make_draft_alert_id(record["source"]["source_id"], record["run"]["run_id"], bare_hash)


def test_rows_join_evidence_assessment_and_action(tmp_path):
    record = _make_canonical_record(tmp_path)
    run_id = record["run"]["run_id"]
    record_id = record["record_id"]
    _write_assessment(tmp_path, evidence_record_id=run_id)
    # The action-log decision is keyed by the REAL draft alert id (source_id |
    # run_id | normalized_hash), NOT by run_id — the register must re-derive it.
    draft_id = _real_draft_alert_id(record)
    assert draft_id.startswith("draft-")
    lookup = _action_lookup(
        {draft_id: [{"decision": "monitor", "created_at": "2026-06-20T02:00:00Z"}]}
    )

    rows = build_change_register_rows(
        user_id=7,
        base_dir=tmp_path,
        action_log_lookup=lookup,
    )

    assert len(rows) == 1
    row = rows[0]
    # Canonical evidence fields
    assert row["date"] == "2026-06-20T00:00:00Z"
    assert row["source_name"] == "DFSA Test Source"
    assert row["regulator_code"] == "DFSA"
    assert row["normalized_hash"] == record["content"]["current_hash"]
    assert row["normalized_hash"].startswith("sha256:")
    assert "changed against the previous" in row["change_summary"]
    # Human assessment fields
    assert row["reviewer"] == "Jane Compliance"
    assert row["impact_level"] == "policy_review"
    assert row["next_action"] == "Route to MLRO for internal review"
    # Action-log decision (scoped to the requesting client)
    assert row["action_decision"] == "monitor"
    assert "impact: policy_review" in row["impact_action"]
    assert "action: monitor" in row["impact_action"]
    # Status + proof reference
    assert row["status"]  # non-empty
    assert record_id in row["proof_reference"]
    assert "evidence-record.json" in row["proof_reference"]


def test_action_decision_joins_only_on_real_draft_id(tmp_path):
    """BLOCK-3: a decision logged ONLY under the real draft id (not run_id or
    record_id) must still surface — proving the register derives the true key."""
    record = _make_canonical_record(tmp_path)
    draft_id = _real_draft_alert_id(record)
    # Keyed exclusively by the draft id — a run_id/record_id lookup would miss it.
    lookup = _action_lookup(
        {draft_id: [{"decision": "act", "created_at": "2026-06-20T09:00:00Z"}]}
    )

    rows = build_change_register_rows(user_id=42, base_dir=tmp_path, action_log_lookup=lookup)

    assert len(rows) == 1
    assert rows[0]["action_decision"] == "act"
    assert "action: act" in rows[0]["impact_action"]


def test_action_decision_real_action_log_round_trip(tmp_path, monkeypatch):
    """Save a real action-log entry under the real draft id via the production
    store, then assert the register (default get_action_log) shows the decision."""
    from app import db as db_module
    from app.alert_actions import get_action_log, save_action_log_entry

    monkeypatch.setattr(db_module, "DB_PATH", str(tmp_path / "actions.db"))
    db_module.ensure_auth_tables()  # creates the alert_action_log table

    record = _make_canonical_record(tmp_path)
    draft_id = _real_draft_alert_id(record)
    saved = save_action_log_entry(user_id=88, alert_id=draft_id, decision="no_action", notes="reviewed")
    assert saved is not None
    assert get_action_log(88, draft_id), "entry must be retrievable under the draft id"

    # No injected lookup → the register uses the real get_action_log store.
    rows = build_change_register_rows(user_id=88, base_dir=tmp_path)
    assert len(rows) == 1
    assert rows[0]["action_decision"] == "no_action"
    assert "action: no_action" in rows[0]["impact_action"]


def test_action_decision_absent_without_user_context(tmp_path):
    record = _make_canonical_record(tmp_path)
    _write_assessment(tmp_path, evidence_record_id=record["run"]["run_id"])

    rows = build_change_register_rows(base_dir=tmp_path)  # no user_id

    assert len(rows) == 1
    assert rows[0]["action_decision"] == ""
    # Impact still surfaces from the assessment even without an action decision.
    assert "impact: policy_review" in rows[0]["impact_action"]


def test_rows_filter_by_source_and_regulator_and_date(tmp_path):
    _make_canonical_record(
        tmp_path,
        source_id="AE-dfsa-a",
        regulator_slug="dfsa",
        regulator="DFSA",
        run_id="run_dfsa_a",
        record_id="evr_dfsa_a",
        timestamp="2026-06-10T00:00:00Z",
    )
    _make_canonical_record(
        tmp_path,
        source_id="AE-vara-b",
        regulator_slug="vara",
        regulator="VARA",
        source_name="VARA Test Source",
        run_id="run_vara_b",
        record_id="evr_vara_b",
        timestamp="2026-07-01T00:00:00Z",
    )

    all_rows = build_change_register_rows(base_dir=tmp_path)
    assert len(all_rows) == 2
    # Rows sorted by date descending → newest first
    assert all_rows[0]["regulator_code"] == "VARA"

    by_source = build_change_register_rows(base_dir=tmp_path, source_id="AE-dfsa-a")
    assert [r["source_id"] for r in by_source] == ["AE-dfsa-a"]

    by_reg = build_change_register_rows(base_dir=tmp_path, regulator="VARA")
    assert [r["regulator_code"] for r in by_reg] == ["VARA"]

    by_range = build_change_register_rows(
        base_dir=tmp_path, date_from="2026-06-01", date_to="2026-06-30"
    )
    assert [r["record_id"] for r in by_range] == ["evr_dfsa_a"]


def test_excluded_source_ids_drops_rows_even_without_source_filter(tmp_path):
    """BLOCK-2: a default (no source_id) export must not surface an excluded
    (e.g. another tenant's private custom) source. Rows for excluded source_ids
    are dropped unconditionally, while shared/official rows survive.
    """
    # custom-A belongs to another tenant; official-X is shared.
    _make_canonical_record(
        tmp_path,
        source_id="custom-A",
        regulator_slug="custom",
        regulator="CUSTOMREG",
        source_name="User A private source",
        run_id="run_custom_a",
        record_id="evr_custom_a",
        timestamp="2026-06-10T00:00:00Z",
    )
    _make_canonical_record(
        tmp_path,
        source_id="official-X",
        regulator_slug="dfsa",
        regulator="DFSA",
        source_name="Shared official source",
        run_id="run_official_x",
        record_id="evr_official_x",
        timestamp="2026-07-01T00:00:00Z",
    )

    # Before exclusion both rows appear.
    all_rows = build_change_register_rows(base_dir=tmp_path)
    assert {r["source_id"] for r in all_rows} == {"custom-A", "official-X"}

    # With custom-A excluded (no source_id filter given), only official-X remains.
    scoped = build_change_register_rows(
        base_dir=tmp_path, excluded_source_ids={"custom-A"}
    )
    assert [r["source_id"] for r in scoped] == ["official-X"]


# ── renderers ───────────────────────────────────────────────────────────────────

def test_csv_render_has_headers_disclaimer_and_data(tmp_path):
    record = _make_canonical_record(tmp_path)
    rows = build_change_register_rows(base_dir=tmp_path)
    meta = {
        "generated_at_utc": "2026-07-11T00:00:00Z",
        "row_count": str(len(rows)),
        "date_from": "(unbounded)",
        "date_to": "(unbounded)",
        "source_id": "(all)",
        "regulator": "(all)",
    }
    text = render_change_register_csv(rows, meta)

    assert LEGAL_DISCLAIMER in text
    assert "Normalized hash" in text
    assert "Regulator" in text
    assert record["source"]["source_name"] in text
    # Parses as CSV once the commented preamble lines are skipped.
    data_lines = [ln for ln in text.splitlines() if not ln.startswith("#")]
    reader = list(csv_reader(data_lines))
    assert reader[0][0] == "Date"
    assert len(reader) == 2  # header + one data row


def csv_reader(lines):
    import csv

    return csv.reader(lines)


def test_html_render_contains_disclaimer_and_table(tmp_path):
    _make_canonical_record(tmp_path)
    rows = build_change_register_rows(base_dir=tmp_path)
    meta = {
        "generated_at_utc": "2026-07-11T00:00:00Z",
        "row_count": str(len(rows)),
        "date_from": "(unbounded)",
        "date_to": "(unbounded)",
        "source_id": "(all)",
        "regulator": "(all)",
    }
    doc = render_change_register_html(rows, meta)

    assert doc.lstrip().startswith("<!doctype html>")
    assert LEGAL_DISCLAIMER in doc
    # Disclaimer appears in both header banner and footer legal boundary.
    assert doc.count(LEGAL_DISCLAIMER) >= 2
    assert "<table>" in doc
    assert "DFSA Test Source" in doc


def test_xlsx_render_is_valid_workbook_with_headers(tmp_path):
    _make_canonical_record(tmp_path)
    rows = build_change_register_rows(base_dir=tmp_path)
    meta = {
        "generated_at_utc": "2026-07-11T00:00:00Z",
        "row_count": str(len(rows)),
        "date_from": "(unbounded)",
        "date_to": "(unbounded)",
        "source_id": "(all)",
        "regulator": "(all)",
    }
    blob = render_change_register_xlsx(rows, meta)

    assert isinstance(blob, bytes)
    assert blob[:2] == b"PK"  # XLSX is a ZIP container

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(blob))
    worksheet = workbook.active
    all_text = "\n".join(
        str(cell.value) for row in worksheet.iter_rows() for cell in row if cell.value is not None
    )
    assert LEGAL_DISCLAIMER in all_text
    assert "Normalized hash" in all_text
    assert "DFSA Test Source" in all_text


# ── forbidden-claims guard ──────────────────────────────────────────────────────

def test_assert_no_forbidden_claims_passes_on_clean_text():
    assert_no_forbidden_claims(f"{LEGAL_DISCLAIMER} Regulator: DFSA. Status: approved.")


@pytest.mark.parametrize("phrase", ["guarantee compliance", "prevent fines", "100% accurate"])
def test_assert_no_forbidden_claims_raises(phrase):
    with pytest.raises(ChangeRegisterError):
        assert_no_forbidden_claims(f"This will {phrase} for you.")


def test_render_blocks_forbidden_claim_in_row_content():
    poisoned = [
        {
            "date": "2026-06-20T00:00:00Z",
            "source_name": "Poison Source",
            "regulator_code": "DFSA",
            "change_summary": "This tool will guarantee compliance for the client.",
            "normalized_hash": "sha256:" + "0" * 64,
            "risk_level": "NOT_SCORED",
            "reviewer": "Reviewer",
            "impact_action": "impact: — / action: —",
            "next_action": "",
            "status": "approved",
            "proof_reference": "evr_x (evidence/x/evidence-record.json)",
        }
    ]
    meta = {
        "generated_at_utc": "2026-07-11T00:00:00Z",
        "row_count": "1",
        "date_from": "(unbounded)",
        "date_to": "(unbounded)",
        "source_id": "(all)",
        "regulator": "(all)",
    }
    with pytest.raises(ChangeRegisterError):
        render_change_register_csv(poisoned, meta)
    with pytest.raises(ChangeRegisterError):
        render_change_register_html(poisoned, meta)
    with pytest.raises(ChangeRegisterError):
        render_change_register_xlsx(poisoned, meta)


@pytest.mark.parametrize(
    "renderer",
    [render_change_register_csv, render_change_register_html, render_change_register_xlsx],
)
def test_forbidden_phrase_in_meta_filter_blocks_all_formats(renderer):
    """MEDIUM-6: a forbidden phrase injected via a filter value (which lands in
    the metadata line, not a data cell) must block ALL three renders equally —
    previously XLSX skipped the meta line and shipped it."""
    meta = {
        "generated_at_utc": "2026-07-11T00:00:00Z",
        "row_count": "0",
        "date_from": "(unbounded)",
        "date_to": "(unbounded)",
        "source_id": "(all)",
        # The regulator filter is free-form and echoed verbatim into the meta line.
        "regulator": "we guarantee compliance",
    }
    with pytest.raises(ChangeRegisterError):
        renderer([], meta)


# ── export (writes files) + empty-but-valid register ────────────────────────────

def test_export_writes_all_three_formats(tmp_path):
    record = _make_canonical_record(tmp_path)
    _write_assessment(tmp_path, evidence_record_id=record["run"]["run_id"])

    result = build_change_register_export(
        user_id=7,
        base_dir=tmp_path,
        action_log_lookup=_action_lookup(
            {record["run"]["run_id"]: [{"decision": "act", "created_at": "2026-06-20T05:00:00Z"}]}
        ),
    )

    assert result["status"] == "ok"
    assert result["row_count"] == 1
    assert result["disclaimer"] == LEGAL_DISCLAIMER
    for key in ("csv_path", "xlsx_path", "html_path"):
        assert key in result["exports"]
        assert (tmp_path / result["exports"][key]).exists()

    html_doc = (tmp_path / result["exports"]["html_path"]).read_text(encoding="utf-8")
    assert LEGAL_DISCLAIMER in html_doc
    assert "act" in html_doc  # the action decision made it into the table


def test_export_single_format(tmp_path):
    _make_canonical_record(tmp_path)
    result = build_change_register_export(base_dir=tmp_path, export_format="csv")
    assert result["status"] == "ok"
    assert set(result["exports"].keys()) == {"csv_path"}


def test_export_unsupported_format_is_error(tmp_path):
    _make_canonical_record(tmp_path)
    result = build_change_register_export(base_dir=tmp_path, export_format="pdf")
    assert result["status"] == "error"


def test_empty_range_yields_empty_but_valid_register(tmp_path):
    # A record exists, but the date range matches nothing.
    _make_canonical_record(tmp_path, timestamp="2026-06-20T00:00:00Z")

    result = build_change_register_export(
        base_dir=tmp_path,
        date_from="2020-01-01",
        date_to="2020-12-31",
    )

    assert result["status"] == "ok"
    assert result["row_count"] == 0
    # Files still written and valid, with headers + disclaimer, zero data rows.
    csv_text = (tmp_path / result["exports"]["csv_path"]).read_text(encoding="utf-8")
    assert LEGAL_DISCLAIMER in csv_text
    assert "Date" in csv_text  # header present
    html_doc = (tmp_path / result["exports"]["html_path"]).read_text(encoding="utf-8")
    assert LEGAL_DISCLAIMER in html_doc
    assert "empty but valid" in html_doc


def test_no_records_at_all_is_empty_but_valid(tmp_path):
    result = build_change_register_export(base_dir=tmp_path)
    assert result["status"] == "ok"
    assert result["row_count"] == 0
    assert (tmp_path / result["exports"]["xlsx_path"]).exists()


# ── record-count cap (DoS bound; reject, never truncate) ───────────────────────

def test_register_rejects_over_record_cap(tmp_path, monkeypatch):
    """Over the record cap → REJECT with a clear message, and write no file."""
    from app import change_register as cr

    over = [
        {"record_id": f"r{i}", "record_path": ""}
        for i in range(cr.MAX_REGISTER_RECORDS + 1)
    ]
    monkeypatch.setattr(cr, "list_canonical_evidence_records", lambda **kw: over)

    # The row builder raises the dedicated too-large error (not the forbidden
    # claims error, and not a silent truncation).
    with pytest.raises(cr.ChangeRegisterTooLargeError):
        cr.build_change_register_rows(base_dir=tmp_path)

    out_dir = tmp_path / "out"
    result = cr.build_change_register_export(base_dir=tmp_path, output_dir=out_dir)
    assert result["status"] == "error"
    assert str(cr.MAX_REGISTER_RECORDS) in result["message"]
    assert "narrow" in result["message"].lower()
    # No partial/truncated register was written to disk.
    assert not out_dir.exists() or not any(out_dir.iterdir())


def test_register_at_record_cap_is_not_rejected(tmp_path, monkeypatch):
    """Exactly at the cap must NOT trip the guard (off-by-one boundary)."""
    from app import change_register as cr

    at_cap = [
        {"record_id": f"r{i}", "record_path": ""}
        for i in range(cr.MAX_REGISTER_RECORDS)
    ]
    monkeypatch.setattr(cr, "list_canonical_evidence_records", lambda **kw: at_cap)

    # Must not raise; the fake summaries load nothing, so rows come back empty —
    # the point is that the cap check itself allowed the export to proceed.
    rows = cr.build_change_register_rows(base_dir=tmp_path)
    assert rows == []


# ── CSV formula-injection neutralization ──────────────────────────────────────

def test_csv_neutralizes_formula_injection_in_authored_fields():
    """Human-authored cells that begin with a formula trigger (= + - @) must be
    prefixed with a single quote so a spreadsheet renders them as text."""
    rows = [{
        "date": "2026-07-13",
        "source_name": "DFSA",
        "regulator_code": "DFSA",
        "change_summary": "=HYPERLINK(\"http://evil\",\"click\")",
        "normalized_hash": "sha256:abc",
        "risk_level": "MEDIUM",
        "reviewer": "@attacker",
        "impact_action": "+1+1",
        "next_action": "-2+3",
        "status": "reviewed",
        "proof_reference": "run-1",
    }]
    meta = {
        "generated_at_utc": "2026-07-13T00:00:00Z", "row_count": "1",
        "date_from": "(unbounded)", "date_to": "(unbounded)",
        "source_id": "(all)", "regulator": "(all)",
    }
    csv_text = render_change_register_csv(rows, meta)
    # Each dangerous cell is quoted; the raw formula never starts a field.
    assert "'=HYPERLINK" in csv_text
    assert "'@attacker" in csv_text
    assert "'+1+1" in csv_text
    assert "'-2+3" in csv_text
    # A leading '=' must never begin a CSV field (which is what Excel executes).
    for line in csv_text.splitlines():
        for field in line.split(","):
            assert not field.lstrip('"').startswith("="), field


def test_xlsx_neutralizes_formula_injection():
    """openpyxl treats a string starting with '=' as a live formula, so the XLSX
    renderer must apply the same guard as CSV."""
    from openpyxl import load_workbook

    rows = [{
        "date": "2026-07-13", "source_name": "DFSA", "regulator_code": "DFSA",
        "change_summary": "=1+1", "normalized_hash": "sha256:abc",
        "risk_level": "LOW", "reviewer": "@x", "impact_action": "monitor",
        "next_action": "-2+3", "status": "reviewed", "proof_reference": "run-1",
    }]
    meta = {
        "generated_at_utc": "2026-07-13T00:00:00Z", "row_count": "1",
        "date_from": "(unbounded)", "date_to": "(unbounded)",
        "source_id": "(all)", "regulator": "(all)",
    }
    data = render_change_register_xlsx(rows, meta)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    formula_cells = [
        c.value for r in ws.iter_rows() for c in r
        if isinstance(c.value, str) and c.value[:1] in ("=", "+", "-", "@")
    ]
    # No cell begins with a raw formula trigger — each dangerous cell was quoted.
    assert formula_cells == [], formula_cells


def test_csv_leaves_safe_cells_unchanged():
    rows = [{
        "date": "2026-07-13", "source_name": "DFSA", "regulator_code": "DFSA",
        "change_summary": "Section 7 amended", "normalized_hash": "sha256:abc",
        "risk_level": "LOW", "reviewer": "Alice", "impact_action": "monitor",
        "next_action": "none", "status": "reviewed", "proof_reference": "run-1",
    }]
    meta = {
        "generated_at_utc": "2026-07-13T00:00:00Z", "row_count": "1",
        "date_from": "(unbounded)", "date_to": "(unbounded)",
        "source_id": "(all)", "regulator": "(all)",
    }
    csv_text = render_change_register_csv(rows, meta)
    assert "Section 7 amended" in csv_text
    assert "'Section" not in csv_text


def test_register_does_not_leak_other_tenant_assessment_when_org_none(tmp_path):
    """Regression: build_change_register_rows must NOT surface another tenant's
    reviewer/impact/next_action into the downloadable register when the caller's
    org is unresolved (org_id=None). Previously `_akw = {} if org_id is None ...`
    dropped the kwarg on None and fell back to _NO_ORG_FILTER, joining ANY org's
    assessment onto the shared official record. Org A stamps a NON-EMPTY org_id,
    so a None caller (scoped to the "" bucket) must see an unassessed row.
    """
    record = _make_canonical_record(tmp_path)
    # Org A (org_id="1") privately assessed this shared official record.
    secret_reviewer = "MLRO-OrgA-PRIVATE"
    secret_action = "Escalate to OrgA board — internal"
    org_a_row = {
        "assessment_id": "assessment-orgA",
        "evidence_record_id": record["run"]["run_id"],
        "impact_level": "policy_review",
        "reviewer_name": secret_reviewer,
        "next_action": secret_action,
        "assessment_status": "acknowledged",
        "reviewed_at": "2026-06-20T01:00:00Z",
        "org_id": "1",
        "legal_disclaimer": LEGAL_DISCLAIMER,
    }
    apath = tmp_path / "data" / "evidence_assessments" / "assessments.jsonl"
    apath.parent.mkdir(parents=True, exist_ok=True)
    apath.write_text(json.dumps(org_a_row, sort_keys=True) + "\n", encoding="utf-8")

    # Unresolved caller (org_id=None) must NOT receive Org A's private note.
    rows_none = build_change_register_rows(base_dir=tmp_path, org_id=None)
    assert len(rows_none) == 1
    blob_none = json.dumps(rows_none)
    assert secret_reviewer not in blob_none
    assert secret_action not in blob_none

    # Control: Org A itself DOES see its own assessment.
    rows_a = build_change_register_rows(base_dir=tmp_path, org_id="1")
    assert secret_reviewer in json.dumps(rows_a)
