"""
RegRadar Enterprise — CIS Regulatory Intelligence Platform
Monolithic production build. Run: python app.py
"""
# ══════════════════════════════════════════════════════════════════
#  BOOTSTRAP — auto-install missing packages before any other import
# ══════════════════════════════════════════════════════════════════
import sys, subprocess, importlib

_DEPS: dict[str, str] = {
    "nicegui":          "nicegui",
    "sqlalchemy":       "sqlalchemy",
    "loguru":           "loguru",
    "openpyxl":         "openpyxl",
    "curl-cffi":        "curl_cffi",
    "openai":           "openai",
    "python-dotenv":    "dotenv",
    "pdfplumber":       "pdfplumber",
    "beautifulsoup4":   "bs4",
    "httpx":            "httpx",
    "instructor":       "instructor",
    "pydantic-settings":"pydantic_settings",
}

def _bootstrap():
    missing = [pkg for pkg, mod in _DEPS.items() if not _try_import(mod)]
    if not missing:
        return
    print(f"[RegRadar] Installing: {', '.join(missing)}", flush=True)
    r = subprocess.run(
        [sys.executable, "-m", "pip", "install", "--quiet", *missing],
        capture_output=True, text=True,
    )
    print("[RegRadar] Bootstrap complete.\n" if r.returncode == 0
          else f"[RegRadar] pip warn: {r.stderr[-300:]}")

def _try_import(name: str) -> bool:
    try: importlib.import_module(name); return True
    except ImportError: return False

_bootstrap()

# ══════════════════════════════════════════════════════════════════
#  IMPORTS
# ══════════════════════════════════════════════════════════════════
import os, asyncio, math, uuid, json, hashlib, html as _html, re
import concurrent.futures
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from loguru import logger
from nicegui import ui, app as _ui_app
from sqlalchemy import func, text

# ── Dirs & logging ─────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
(BASE_DIR / "logs").mkdir(exist_ok=True)
(BASE_DIR / "data").mkdir(exist_ok=True)
ENV_PATH = BASE_DIR.parent.parent.parent / ".env"

logger.remove()
logger.add(sys.stdout, colorize=True,
           format="<green>{time:HH:mm:ss}</green> | <level>{level:<7}</level> | {message}")
logger.add(BASE_DIR / "logs" / "app.log", rotation="10 MB",
           format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}")

# ── DB layer ───────────────────────────────────────────────────────
from database import (
    init_db as _init_db, RegulationRecord, Regulator,
    SessionLocal, engine, reg_repo, reg_repo_r, job_repo, DEFAULT_REGULATORS,
)
from app.application.pipeline import process_regulator as _real_scrape
from app.core.domain.enums import Jurisdiction as _JurEnum

try:
    from curl_cffi import requests as _cffi_req
    _HAS_CURL_CFFI = True
except ImportError:
    _HAS_CURL_CFFI = False

# ══════════════════════════════════════════════════════════════════
#  DB MIGRATION — add Enterprise columns & tables
# ══════════════════════════════════════════════════════════════════

def _migrate_db():
    with engine.connect() as c:
        existing = {r[1] for r in c.execute(text("PRAGMA table_info(regulations)"))}
        for col, defn in [
            ("ai_analysis",   "TEXT DEFAULT '{}'"),
            ("source_type",   "VARCHAR(20) DEFAULT 'WEB'"),
            ("content_hash",  "TEXT DEFAULT NULL"),
        ]:
            if col not in existing:
                c.execute(text(f"ALTER TABLE regulations ADD COLUMN {col} {defn}"))
        c.execute(text("""
            CREATE TABLE IF NOT EXISTS company_policies (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                name    TEXT NOT NULL,
                content TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )"""))
        c.execute(text("""
            CREATE TABLE IF NOT EXISTS webhook_settings (
                key   TEXT PRIMARY KEY,
                value TEXT DEFAULT ''
            )"""))
        # Журнал аудита: каждое изменение статуса документа
        c.execute(text("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                doc_id     INTEGER NOT NULL,
                old_status TEXT NOT NULL,
                new_status TEXT NOT NULL,
                changed_at TEXT NOT NULL
            )"""))
        # Fix: VALIDATED records with empty ai_analysis are contradictory — reset to ACTIVE
        c.execute(text(
            "UPDATE regulations SET status='ACTIVE' "
            "WHERE status='VALIDATED' AND (ai_analysis IS NULL OR ai_analysis='{}')"
        ))
        # Migrate regulator URLs to correct values + add strategy column
        reg_cols = {r[1] for r in c.execute(text("PRAGMA table_info(regulators)"))}
        if "strategy" not in reg_cols:
            c.execute(text("ALTER TABLE regulators ADD COLUMN strategy VARCHAR(20) DEFAULT 'STATIC'"))
        # Fix broken URLs
        c.execute(text(
            "UPDATE regulators SET base_url='https://nationalbank.kz/ru/link/normativnaya-pravovaya-baza',"
            "strategy='JS_RENDER' WHERE jurisdiction='KZ' AND base_url LIKE '%normativy%'"
        ))
        c.execute(text(
            "UPDATE regulators SET strategy='JS_RENDER' WHERE jurisdiction='BY' AND base_url LIKE '%nbrb.by%'"
        ))
        # CBR requires RU IP — AUTO tries STATIC → SITEMAP → JS_RENDER
        c.execute(text(
            "UPDATE regulators SET strategy='AUTO' WHERE jurisdiction='RU' AND base_url LIKE '%cbr.ru%'"
        ))
        c.commit()

# ══════════════════════════════════════════════════════════════════
#  MULTI-AGENT AI ANALYSIS  (Legal → Risk → Action)
# ══════════════════════════════════════════════════════════════════

def _ai_analyze(raw_text: str, jurisdiction: str) -> dict:
    from app.infrastructure.ai.analyzer import run_multi_agent
    return run_multi_agent(raw_text, jurisdiction)


def _check_policies(reg_text: str) -> str:
    """Compare regulation text against stored company policies via OpenAI."""
    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        return ""
    with engine.connect() as c:
        rows = c.execute(text("SELECT name, content FROM company_policies")).fetchall()
    if not rows:
        return ""
    policies_block = "\n\n".join(
        f"Policy '{r[0]}':\n{r[1][:600]}" for r in rows[:4]
    )
    from openai import OpenAI
    resp = OpenAI(api_key=api_key).chat.completions.create(
        model="gpt-4o-mini", max_tokens=180, temperature=0.0,
        messages=[
            {"role": "system", "content":
                "You are a compliance officer. Identify specific conflicts between the new regulation and "
                "internal policies. If conflict found, name the policy and section precisely (1-2 sentences). "
                "If no conflict, reply exactly: NO_CONFLICT"},
            {"role": "user", "content":
                f"REGULATION:\n{reg_text[:1800]}\n\nINTERNAL POLICIES:\n{policies_block}"},
        ],
    ).choices[0].message.content or ""
    return "" if "NO_CONFLICT" in resp.upper() else resp.strip()

# ══════════════════════════════════════════════════════════════════
#  WEBHOOK SENDERS
# ══════════════════════════════════════════════════════════════════

def _wh_get(key: str) -> str:
    try:
        with engine.connect() as c:
            row = c.execute(text("SELECT value FROM webhook_settings WHERE key=:k"),
                            {"k": key}).fetchone()
        return row[0] if row else ""
    except Exception:
        return ""

def _wh_set(key: str, value: str):
    with engine.connect() as c:
        c.execute(text("INSERT OR REPLACE INTO webhook_settings(key,value) VALUES(:k,:v)"),
                  {"k": key, "v": value})
        c.commit()

async def _tg_send(token: str, chat_id: str, msg: str) -> bool:
    if not token or not chat_id:
        return False
    try:
        import httpx
        async with httpx.AsyncClient(timeout=10) as h:
            r = await h.post(
                f"https://api.telegram.org/bot{token}/sendMessage",
                json={"chat_id": chat_id, "text": msg, "parse_mode": "HTML"},
            )
        return r.status_code == 200
    except Exception as e:
        logger.warning("Telegram alert failed: {}", e); return False

async def _slack_send(url: str, msg: str) -> bool:
    if not url:
        return False
    try:
        import httpx
        async with httpx.AsyncClient(timeout=10) as h:
            r = await h.post(url, json={"text": msg})
        return r.status_code == 200
    except Exception as e:
        logger.warning("Slack alert failed: {}", e); return False

async def _jira_create(endpoint: str, token: str, project: str,
                        title: str, desc: str) -> str:
    if not endpoint or not token:
        return ""
    try:
        import httpx
        async with httpx.AsyncClient(timeout=15) as h:
            r = await h.post(
                f"{endpoint.rstrip('/')}/rest/api/2/issue",
                headers={"Authorization": f"Basic {token}",
                         "Content-Type": "application/json"},
                json={"fields": {
                    "project": {"key": project or "COMP"},
                    "summary": f"[RegRadar] {title[:80]}",
                    "description": desc[:1000],
                    "issuetype": {"name": "Task"},
                }},
            )
        return (r.json().get("key", "") if r.status_code in (200, 201) else "")
    except Exception as e:
        logger.warning("Jira create failed: {}", e); return ""

async def _dispatch_alerts(new_critical: list[dict]):
    if not new_critical:
        return
    tok   = _wh_get("tg_token")
    cid   = _wh_get("tg_chat_id")
    slack = _wh_get("slack_url")
    jira_ep  = _wh_get("jira_endpoint")
    jira_tok = _wh_get("jira_token")
    jira_prj = _wh_get("jira_project")
    for reg in new_critical:
        ai  = reg.get("ai_data", {})
        act = "\n".join(f"• {s}" for s in ai.get("action", []))
        tg_msg = (
            f"🚨 <b>CRITICAL — RegRadar</b>\n\n"
            f"📍 <b>{reg['jurisdiction']}</b> — {reg['title'][:90]}\n\n"
            f"💰 Max fine: {ai.get('fines', 'N/A')}\n\n"
            f"📋 <b>Action Plan:</b>\n{act[:500]}"
        )
        slack_msg = (f"[RegRadar CRITICAL] {reg['jurisdiction']} — "
                     f"{reg['title'][:80]}  |  Fine: {ai.get('fines', 'N/A')}")
        await _tg_send(tok, cid, tg_msg)
        await _slack_send(slack, slack_msg)
        if jira_ep:
            await _jira_create(jira_ep, jira_tok, jira_prj, reg["title"], act)

# ══════════════════════════════════════════════════════════════════
#  CURL_CFFI FALLBACK SCRAPER
# ══════════════════════════════════════════════════════════════════

def _fallback_scrape(base_url: str, jurisdiction: str, link_pattern: str) -> int:
    if not _HAS_CURL_CFFI:
        return 0
    import io, pdfplumber
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse
    from app.infrastructure.ai.extractor import extract_regulation

    sess = _cffi_req.Session(impersonate="chrome120")
    resp = sess.get(base_url, timeout=30)
    if resp.status_code >= 400:
        return 0
    soup  = BeautifulSoup(resp.text, "html.parser")
    base  = urlparse(base_url)
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if link_pattern.lower() not in href.lower():
            continue
        links.append(href if href.startswith("http")
                     else f"{base.scheme}://{base.netloc}{href}")
    saved = 0
    for url in links[:8]:
        try:
            pr = sess.get(url, timeout=30)
            if pr.status_code >= 400:
                continue
            with pdfplumber.open(io.BytesIO(pr.content)) as pdf:
                raw = " ".join(p.extract_text() or "" for p in pdf.pages[:3])
            try:
                je = _JurEnum(jurisdiction)
            except ValueError:
                je = None
            m = extract_regulation(raw, je, url)
            rec = {
                "title":            m.title,
                "jurisdiction":     jurisdiction,
                "publication_date": str(m.publication_date or ""),
                "effective_date":   str(m.effective_date or ""),
                "summary":          m.summary,
                "industries":       ",".join(m.industries),
                "critical_level":   (m.critical_level.value
                                     if hasattr(m.critical_level, "value")
                                     else str(m.critical_level)),
                "source_url":       url,
                "confidence":       m.confidence,
                "status":           "VALIDATED" if m.confidence >= 0.85 else "HUMAN_REVIEW",
                "extracted_at":     datetime.utcnow(),
            }
            if reg_repo.upsert(rec):
                saved += 1
        except Exception:
            continue
    return saved

# ══════════════════════════════════════════════════════════════════
#  SCRAPE SESSION
# ══════════════════════════════════════════════════════════════════

@dataclass
class RegRun:
    name:      str
    jur:       str
    status:    str = "RUNNING"
    found:     int = 0
    saved:     int = 0
    bytes_dl:  int = 0
    method:    str = "requests"
    error:     str = ""

@dataclass
class ScrapeSession:
    sid:       str            = field(default_factory=lambda: uuid.uuid4().hex[:8].upper())
    started:   datetime       = field(default_factory=datetime.now)
    status:    str            = "RUNNING"
    runs:      list[RegRun]   = field(default_factory=list)
    finished:  Optional[datetime] = None

    @property
    def found(self):  return sum(r.found  for r in self.runs)
    @property
    def saved(self):  return sum(r.saved  for r in self.runs)
    @property
    def bytes_dl(self): return sum(r.bytes_dl for r in self.runs)
    @property
    def error(self):  return "; ".join(r.error for r in self.runs if r.error)
    @property
    def duration(self):
        e = self.finished or datetime.now()
        return f"{int((e - self.started).total_seconds())}s"

# ══════════════════════════════════════════════════════════════════
#  SEED DATA — 10 deep Enterprise records
# ══════════════════════════════════════════════════════════════════

def _j(d): return json.dumps(d, ensure_ascii=False)

_SAMPLES = [
  dict(
    title="Положение ЦБ РФ №894-П: RT-мониторинг P2P-транзакций и KYC в шлюзах",
    jurisdiction="RU", critical_level="CRITICAL",
    publication_date="2026-01-10", effective_date="2026-06-01",
    summary="ЦБ РФ ввёл обязательный RT-мониторинг P2P-платежей >600 тыс. руб./сутки.",
    industries="banking,payments,AML",
    source_url="https://cbr.ru/reg/894p-rt-2026",
    confidence=0.97, status="VALIDATED", source_type="WEB",
    extracted_at=datetime(2026, 1, 15, 9, 0),
    ai_analysis=_j({
      "legal": "Согласно ст. 7 ФЗ-115 «О ПОД/ФТ», п. 1.1 введён критерий подозрительности P2P-операций свыше 600 тыс. руб./сут. при отсутствии KYC. Срок внедрения RT-систем мониторинга — 01.06.2026. Ответственность банка — административная по ст. 15.27 КоАП.",
      "risk": "Несоответствие влечёт штраф до 1% от квартального оборота. Для среднего банка (оборот 100 млрд руб./кв.) — до 250 млн руб. Риск отзыва лицензии при повторном нарушении.",
      "fines": "$2,750,000",
      "urgency_days": 60,
      "action": [
        "Провести аудит P2P-транзакций за Q4 2025 до 01.03.2026",
        "Внедрить RT-мониторинг API с порогом 600 тыс. руб./сут. до 01.05.2026",
        "Обновить AML-политику раздел 7.3: снизить лимит мониторинга",
        "Назначить DPO-офицера ответственным за новый регламент",
        "Провести тренинг комплаенс-команды по новым критериям KYC",
      ],
      "internal_conflict": "Конфликт с Политикой AML/KYC раздел 4.2: текущий порог мониторинга 1 млн руб. — требуется снизить до 600 тыс. руб. до 01.05.2026.",
    }),
  ),
  dict(
    title="НК Азербайджана: НДС 18% у источника для иностранных B2B SaaS-провайдеров",
    jurisdiction="AZ", critical_level="CRITICAL",
    publication_date="2026-01-15", effective_date="2026-04-15",
    summary="Иностранные SaaS и цифровые B2B-сервисы обязаны регистрироваться в VOEN.",
    industries="taxation,fintech,SaaS",
    source_url="https://www.nalog.gov.az/reg/voen-saas-2026",
    confidence=0.94, status="VALIDATED", source_type="WEB",
    extracted_at=datetime(2026, 1, 22, 11, 0),
    ai_analysis=_j({
      "legal": "Поправки в ст. 169 НК Азербайджана: иностранные SaaS-компании, оказывающие B2B-услуги юрлицам АР, обязаны встать на учёт в VOEN и удерживать НДС 18% у источника. Ретроактивное действие с 01.01.2026. Срок регистрации — до 15.04.2026.",
      "risk": "Штрафы за уклонение увеличены в 3 раза (до 300% суммы НДС). Блокировка платежей от AZ-банков без VOEN. Риск потери рынка Азербайджана для нерезидентов.",
      "fines": "$180,000",
      "urgency_days": 45,
      "action": [
        "Зарегистрироваться в VOEN через портал e-taxes.gov.az до 01.04.2026",
        "Настроить выставление НДС-счетов для AZ-клиентов",
        "Провести ретроактивный расчёт НДС с 01.01.2026",
        "Обновить договорные шаблоны с AZ-партнёрами",
      ],
      "internal_conflict": "Конфликт с Политикой ценообразования раздел 2.1: тарифы для AZ-клиентов не включают НДС — необходима срочная актуализация прайс-листов.",
    }),
  ),
  dict(
    title="Постановление АРРФР РК №157: AML-скоринг крипто-кошельков в МФЦА (Travel Rule)",
    jurisdiction="KZ", critical_level="HIGH",
    publication_date="2026-02-01", effective_date="2026-05-01",
    summary="Все банки МФЦА обязаны внедрить Travel Rule-совместимый AML-скоринг в реальном времени.",
    industries="banking,crypto,AML",
    source_url="https://www.gov.kz/reg/arrfr-157-2026",
    confidence=0.91, status="VALIDATED", source_type="WEB",
    extracted_at=datetime(2026, 2, 5, 14, 0),
    ai_analysis=_j({
      "legal": "Постановление АРРФР №157 обязывает банки в периметре МФЦА внедрить VASP-идентификацию по Travel Rule (FATF Rec. 16) для транзакций крипто-активов от $1000. Ежеквартальные отчёты об операционных рисках — новая форма 7-ОР-КА.",
      "risk": "Непредоставление отчётности — штраф 500 МРП (~$3,400) за каждый случай. Системные нарушения Travel Rule — приостановление VASP-лицензии МФЦА.",
      "fines": "$85,000",
      "urgency_days": 90,
      "action": [
        "Интегрировать Chainalysis/Elliptic API для VASP-скоринга до 01.04.2026",
        "Разработать форму 7-ОР-КА и регламент ежеквартальной отчётности",
        "Протестировать Travel Rule протокол с тремя топ-VASP-партнёрами",
        "Обновить KYC-анкеты для клиентов с крипто-операциями",
        "Направить тест-отчёт в АРРФР до 15.04.2026",
      ],
      "internal_conflict": "",
    }),
  ),
  dict(
    title="Постановление НБРБ №72: снижение лимитов валютных переводов физлиц до $10k/мес.",
    jurisdiction="BY", critical_level="MEDIUM",
    publication_date="2026-02-20", effective_date="2026-03-20",
    summary="Ежемесячный лимит переводов USD/EUR снижен с $50k до $10k без подтверждения источника.",
    industries="banking,forex,payments",
    source_url="https://www.nbrb.by/reg/72-fx-2026",
    confidence=0.88, status="VALIDATED", source_type="WEB",
    extracted_at=datetime(2026, 2, 25, 10, 0),
    ai_analysis=_j({
      "legal": "Постановление НБРБ №72 устанавливает новый лимит переводов для физлиц в иностранной валюте: не более $10 000 в месяц без декларирования источника средств. Исключения для PayPal, Wise, Stripe отменены. Банки обязаны запрашивать декларацию при превышении.",
      "risk": "Нарушение контроля — предупреждение НБРБ, при повторном нарушении — штраф до 500 БВ (~$6,000). Операционный риск: задержки транзакций клиентов выше лимита.",
      "fines": "$6,000",
      "urgency_days": 28,
      "action": [
        "Обновить параметры ABS-системы: лимит 10k USD/EUR в месяц на клиента",
        "Разработать форму декларации источника средств",
        "Уведомить клиентов с типовыми переводами >$10k (рассылка)",
        "Обучить фронт-офис новым требованиям НБРБ",
      ],
      "internal_conflict": "Конфликт с Процедурой обслуживания VIP-клиентов раздел 3: автоматическое одобрение переводов до $50k — требуется отключить до 15.03.2026.",
    }),
  ),
  dict(
    title="Указ Президента РУз №ПП-321: локализация данных финтех-платформ до 01.01.2027",
    jurisdiction="UZ", critical_level="HIGH",
    publication_date="2025-12-01", effective_date="2027-01-01",
    summary="МФО и платёжные системы с данными граждан РУз обязаны перенести серверы в страну.",
    industries="fintech,data_protection,payments",
    source_url="https://lex.uz/reg/pp-321-localisation-2027",
    confidence=0.93, status="VALIDATED", source_type="WEB",
    extracted_at=datetime(2026, 3, 1, 8, 0),
    ai_analysis=_j({
      "legal": "Указ ПП-321 обязывает все платёжные системы, МФО и операторов персональных данных граждан РУз перенести серверы обработки на территорию страны до 01.01.2027. Надзор возложен на УзКиберЦентр. Штраф за нарушение — до 500 МРЗП.",
      "risk": "500 МРЗП ≈ $25,000. Блокировка работы в РУз при несоответствии. Миграция облачной инфраструктуры на УЦО (~2-6 месяцев работ).",
      "fines": "$25,000",
      "urgency_days": 270,
      "action": [
        "Провести аудит инфраструктуры: какие сервисы обрабатывают данные граждан РУз",
        "Выбрать аккредитованный узбекский ЦОД (UzInfoCom, TBC) до 01.06.2026",
        "Разработать план миграции с временными рамками Q3-Q4 2026",
        "Подать уведомление в УзКиберЦентр о начале работ",
        "Провести юридический аудит договоров с клиентами РУз",
      ],
      "internal_conflict": "",
    }),
  ),
  dict(
    title="Приказ Минфина РФ №42н: новые XSD-схемы деклараций НДС v5.09 и прибыль v3.21",
    jurisdiction="RU", critical_level="MEDIUM",
    publication_date="2026-03-08", effective_date="2026-10-01",
    summary="Переход на новые XML-форматы деклараций НДС и налога на прибыль обязателен с 01.10.2026.",
    industries="taxation,accounting,ERP",
    source_url="https://minfin.gov.ru/reg/prikaz_42n_2026.pdf",
    confidence=0.99, status="VALIDATED", source_type="WEB",
    extracted_at=datetime(2026, 3, 15, 12, 0),
    ai_analysis=_j({
      "legal": "Приказ Минфина №42н утверждает XSD-схемы НДС v5.09 и налога на прибыль v3.21. Старые форматы будут отклоняться ФНС-шлюзом автоматически с 01.10.2026. Налогоплательщики обязаны перейти на новые форматы самостоятельно через провайдеров ЭДО.",
      "risk": "Технические сбои при сдаче отчётности — штраф 5% от суммы налога за каждый месяц просрочки (ст. 119 НК РФ). Для крупного налогоплательщика — до десятков млн руб.",
      "fines": "$220,000",
      "urgency_days": 180,
      "action": [
        "Запросить у поставщика ERP/ЭДО сроки обновления до формата НДС v5.09",
        "Протестировать тестовую выгрузку декларации в ФНС sandbox до 01.07.2026",
        "Обновить XSLT-трансформации в учётной системе",
        "Провести нагрузочный тест подачи на UAT-среде до 01.09.2026",
      ],
      "internal_conflict": "",
    }),
  ),
  dict(
    title="CBAR АР: требования Базель III к достаточности капитала CET1 ≥ 9% с 2027",
    jurisdiction="AZ", critical_level="HIGH",
    publication_date="2026-04-01", effective_date="2027-01-01",
    summary="ЦБА Азербайджана вводит требования Базель III: CET1 не менее 9%, LCR ≥ 100%.",
    industries="banking,capital_adequacy,Basel",
    source_url="https://www.cbar.az/reg/basel3-2027",
    confidence=0.89, status="VALIDATED", source_type="WEB",
    extracted_at=datetime(2026, 4, 5, 16, 0),
    ai_analysis=_j({
      "legal": "Инструкция ЦБА №15 вводит Basel III-требования: коэффициент CET1 ≥ 9%, LCR ≥ 100%, NSFR ≥ 100% для банков с активами >500 млн манат. Переходный период — до 01.01.2027. Ежеквартальная отчётность по форме Basel-Q.",
      "risk": "Нарушение CET1 — ограничение на выплату дивидендов и бонусов. При CET1 < 6% — регуляторное вмешательство ЦБА вплоть до назначения временной администрации.",
      "fines": "$500,000",
      "urgency_days": 250,
      "action": [
        "Рассчитать текущий CET1 по методологии Basel III (Q2 2026)",
        "Разработать план докапитализации при дефиците CET1",
        "Внедрить расчёт LCR и NSFR в ежедневную отчётность казначейства",
        "Подготовить первый отчёт Basel-Q до 01.07.2026 (тест)",
        "Провести стресс-тест портфеля на сценарий роста NPL +3%",
      ],
      "internal_conflict": "",
    }),
  ),
  dict(
    title="АРРФР РК: обязательный SWIFT sanctions-скоринг для корреспондентских счетов МФЦА",
    jurisdiction="KZ", critical_level="CRITICAL",
    publication_date="2026-04-10", effective_date="2026-07-01",
    summary="Банки МФЦА обязаны внедрить RT-скрининг SWIFT-транзакций против санкционных списков.",
    industries="banking,sanctions,AML,SWIFT",
    source_url="https://www.gov.kz/reg/swift-sanctions-2026",
    confidence=0.95, status="VALIDATED", source_type="WEB",
    extracted_at=datetime(2026, 4, 12, 10, 0),
    ai_analysis=_j({
      "legal": "Постановление АРРФР: обязательный RT-скрининг всех SWIFT-транзакций от $500 против списков SDN (OFAC), EU Consolidated List, UN Sanctions. Система должна быть сертифицирована АРРФР. Срок — 01.07.2026.",
      "risk": "Проведение санкционной транзакции — штраф $1 млн + отзыв корреспондентских отношений с US/EU банками. Репутационный и регуляторный риск высокий.",
      "fines": "$1,000,000",
      "urgency_days": 75,
      "action": [
        "Выбрать и интегрировать сертифицированную систему sanctions-скрининга (Fircosoft/Dow Jones) до 01.05.2026",
        "Провести GAP-анализ текущего скрининга SWIFT vs. новые требования",
        "Настроить RT-алёрты для compliance-офицера при hit-е санкционного списка",
        "Получить сертификацию АРРФР до 15.06.2026",
        "Протестировать 1000 тестовых транзакций на UAT",
      ],
      "internal_conflict": "Конфликт с Процедурой обработки платежей раздел 5.3: текущий скрининг выполняется batch-режимом раз в день — требуется переход на RT.",
    }),
  ),
  # ── Telegram Intel records ────────────────────────────────────────
  dict(
    title="[ПРОЕКТ] ЦБ РФ: лицензирование криптовалютных бирж — утечка внутреннего проекта",
    jurisdiction="RU", critical_level="CRITICAL",
    publication_date="2026-05-01", effective_date="2026-12-01",
    summary="Инсайд из Telegram: ЦБ РФ готовит законопроект об обязательном лицензировании CEX.",
    industries="crypto,licensing,fintech",
    source_url="https://t.me/cbr_insider_channel/482",
    confidence=0.72, status="HUMAN_REVIEW", source_type="TELEGRAM_INTEL",
    extracted_at=datetime(2026, 5, 5, 8, 30),
    ai_analysis=_j({
      "legal": "По данным Telegram-канала @cbr_insider, ЦБ РФ направил в Минфин проект поправок в ФЗ-259 о цифровых финансовых активах, вводящих обязательное лицензирование криптовалютных бирж с капиталом ≥ 300 млн руб. Официальная публикация ожидается Q3 2026. Статус: ИНСАЙД — не официальный документ.",
      "risk": "При принятии закона: все действующие CEX без лицензии будут обязаны приостановить работу в РФ. Высокий риск для участников рынка, работающих в серой зоне.",
      "fines": "$3,300,000",
      "urgency_days": 120,
      "action": [
        "Мониторить официальный портал regulation.gov.ru на появление законопроекта",
        "Провести pre-compliance assessment: соответствие требованиям к капиталу и AML",
        "Подготовить черновой комплект документов для будущей лицензии ЦБ",
        "Нанять юридического советника по ФЗ-259 и новым крипто-поправкам",
      ],
      "internal_conflict": "",
    }),
  ),
  dict(
    title="[ПРОЕКТ] НБРБ: цифровой рубль — драфт концепции интеграции с платёжными системами",
    jurisdiction="BY", critical_level="HIGH",
    publication_date="2026-05-03", effective_date="2026-09-01",
    summary="Telegram-инсайд: НБРБ рассматривает обязательное подключение банков к цифровому рублю.",
    industries="CBDC,banking,payments",
    source_url="https://t.me/nbrb_analysis/211",
    confidence=0.68, status="HUMAN_REVIEW", source_type="TELEGRAM_INTEL",
    extracted_at=datetime(2026, 5, 8, 9, 0),
    ai_analysis=_j({
      "legal": "Telegram-канал @nbrb_analysis опубликовал предположительно внутренний документ НБРБ о концепции интеграции платёжной системы цифрового рубля (ЦБДС). Банки должны будут открыть технические счета в НБРБ для ЦБДС-расчётов. Статус: ПРОЕКТ, не официальный документ.",
      "risk": "При принятии: обязательная техническая интеграция API ЦБДС (~3-6 месяцев разработки). Операционные риски трансформации платёжной архитектуры.",
      "fines": "$50,000",
      "urgency_days": 100,
      "action": [
        "Подписаться на официальные рассылки НБРБ для отслеживания официальной версии",
        "Провести предварительный технический анализ интеграции ЦБДС-API",
        "Выделить команду R&D для PoC цифрового рубля",
      ],
      "internal_conflict": "",
    }),
  ),
]


def _seed_samples():
    with SessionLocal() as db:
        if db.query(func.count(RegulationRecord.id)).scalar() > 0:
            return
    with engine.connect() as c:
        for s in _SAMPLES:
            h = hashlib.sha256(s["source_url"].encode()).hexdigest()
            row = c.execute(text("SELECT id FROM regulations WHERE url_hash=:h"),
                            {"h": h}).fetchone()
            if row:
                continue
            c.execute(text("""
                INSERT INTO regulations
                  (title,jurisdiction,critical_level,publication_date,effective_date,
                   summary,industries,source_url,confidence,status,extracted_at,url_hash,
                   ai_analysis,source_type)
                VALUES
                  (:title,:jurisdiction,:critical_level,:publication_date,:effective_date,
                   :summary,:industries,:source_url,:confidence,:status,:extracted_at,:url_hash,
                   :ai_analysis,:source_type)
            """), {
                "title":            s["title"],
                "jurisdiction":     s["jurisdiction"],
                "critical_level":   s["critical_level"],
                "publication_date": s.get("publication_date",""),
                "effective_date":   s.get("effective_date",""),
                "summary":          s["summary"],
                "industries":       s.get("industries",""),
                "source_url":       s["source_url"],
                "confidence":       s.get("confidence", 0.9),
                "status":           s.get("status","VALIDATED"),
                "extracted_at":     s.get("extracted_at", datetime.utcnow()),
                "url_hash":         h,
                "ai_analysis":      json.dumps(s.get("ai_analysis", {}), ensure_ascii=False),
                "source_type":      s.get("source_type","WEB"),
            })
        c.commit()
    logger.success("Seeded {} Enterprise records", len(_SAMPLES))


def _purge_stale_records():
    """Delete records extracted before 2026 — keeps demo trend chart clean."""
    with engine.connect() as c:
        deleted = c.execute(
            text("DELETE FROM regulations WHERE strftime('%Y', extracted_at) < '2026'")
        ).rowcount
        c.commit()
    if deleted:
        print(f"[init] Purged {deleted} pre-2026 records from DB")


def init_app_db():
    _init_db()
    _migrate_db()
    _purge_stale_records()
    reg_repo_r.seed(DEFAULT_REGULATORS)
    _seed_samples()


# ── REST endpoint for status workflow buttons ─────────────────────
_ALLOWED_STATUSES = {"ACTIVE", "ACKNOWLEDGED", "IN_PROGRESS", "DONE", "HUMAN_REVIEW"}

# ── Enterprise Security Stack ─────────────────────────────────────
# Порядок add_middleware: LIFO при выполнении — последний добавленный
# является outermost и обрабатывает запрос первым.
# Добавляем JWT первым → он innermost. SecurityHeaders последним → outermost.
# Итоговый порядок выполнения: SecurityHeaders → RateLimit → JWT → handler.
from app.infrastructure.security.middleware import (
    JWTAuthMiddleware,
    RateLimitMiddleware,
    SecurityHeadersMiddleware,
)

_ui_app.add_middleware(JWTAuthMiddleware)
_ui_app.add_middleware(RateLimitMiddleware)
_ui_app.add_middleware(SecurityHeadersMiddleware)

# ── Монтирование API v1 router ────────────────────────────────────
from app.api.v1.router import api_router as _v1_router
_ui_app.include_router(_v1_router, prefix="/api/v1")


@_ui_app.get("/api/reg/status")
async def _api_set_status(id: int, status: str):
    if status not in _ALLOWED_STATUSES:
        return {"ok": False, "error": "invalid status"}
    with engine.connect() as c:
        # Получаем старый статус для журнала аудита
        row = c.execute(text("SELECT status FROM regulations WHERE id=:i"), {"i": id}).fetchone()
        old_status = row[0] if row else "UNKNOWN"
        c.execute(text("UPDATE regulations SET status=:s WHERE id=:i"), {"s": status, "i": id})
        # Записываем событие в audit_log
        c.execute(text("""
            INSERT INTO audit_log (doc_id, old_status, new_status, changed_at)
            VALUES (:doc_id, :old, :new, :ts)
        """), {
            "doc_id": id,
            "old":    old_status,
            "new":    status,
            "ts":     datetime.utcnow().isoformat(),
        })
        c.commit()
    return {"ok": True}


@_ui_app.get("/api/report/summary")
async def _api_executive_summary(days: int = 7):
    """
    Генерирует Executive Summary для руководства.
    Параметр days — период анализа (по умолчанию 7 дней).
    """
    since = (datetime.utcnow() - timedelta(days=days)).isoformat()
    with engine.connect() as c:
        # Документы за период
        new_docs = c.execute(text(
            "SELECT id, title, critical_level, jurisdiction, "
            "COALESCE(ai_analysis,'{}') ai_analysis "
            "FROM regulations WHERE extracted_at >= :since"
        ), {"since": since}).fetchall()

        # Все активные (не DONE) документы — для расчёта exposure
        active_docs = c.execute(text(
            "SELECT COALESCE(ai_analysis,'{}') ai_analysis "
            "FROM regulations WHERE status NOT IN ('DONE')"
        )).fetchall()

        # Последние события аудита
        audit_events = c.execute(text(
            "SELECT doc_id, old_status, new_status, changed_at "
            "FROM audit_log WHERE changed_at >= :since "
            "ORDER BY changed_at DESC LIMIT 50"
        ), {"since": since}).fetchall()

    # Считаем риски по новым документам
    critical_new = sum(1 for r in new_docs if r.critical_level == "CRITICAL")
    high_new     = sum(1 for r in new_docs if r.critical_level == "HIGH")

    # Суммарный exposure по активным задачам
    total_exposure = 0
    for r in active_docs:
        try:
            ai = json.loads(r.ai_analysis or "{}")
            fv = ai.get("fines")
            if fv and fv not in ("", "N/A", "—"):
                raw = str(fv).replace("$", "").replace(",", "").strip()
                total_exposure += int(float(raw))
        except Exception:
            pass

    resolved = sum(1 for e in audit_events if e.new_status == "DONE")
    hours_saved = len(new_docs) * 1.5

    # Форматируем отчёт
    lines = [
        f"REGRADA EXECUTIVE SUMMARY",
        f"Period: last {days} days  |  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC",
        "=" * 60,
        "",
        f"NEW REGULATORY CHANGES DETECTED: {len(new_docs)}",
        f"  • Critical:  {critical_new}",
        f"  • High risk: {high_new}",
        f"  • Other:     {len(new_docs) - critical_new - high_new}",
        "",
        f"TOTAL ACTIVE RISK EXPOSURE:  ${total_exposure:,}",
        f"RESOLVED IN PERIOD:          {resolved} documents",
        f"HOURS SAVED BY AI:           {hours_saved:.0f}h  ({len(new_docs)} docs × 1.5h)",
        "",
        "TOP CRITICAL CHANGES:",
    ]
    for r in new_docs:
        if r.critical_level == "CRITICAL":
            lines.append(f"  [{r.jurisdiction}] {r.title[:80]}")

    if not any(r.critical_level == "CRITICAL" for r in new_docs):
        lines.append("  No critical changes in this period.")

    lines += [
        "",
        "RECENT STATUS CHANGES (AUDIT):",
    ]
    for e in audit_events[:10]:
        lines.append(
            f"  {e.changed_at[:16]}  doc#{e.doc_id}  "
            f"{e.old_status} → {e.new_status}"
        )
    if not audit_events:
        lines.append("  No status changes in this period.")

    lines += ["", "=" * 60, "Generated by RegRadar AI · Confidential"]
    return {"ok": True, "report": "\n".join(lines), "period_days": days}


# ══════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════

FLAGS       = {"RU":"🇷🇺","KZ":"🇰🇿","AZ":"🇦🇿","BY":"🇧🇾","UZ":"🇺🇿"}
PIE_COLORS  = ["#3b82f6","#10b981","#f59e0b","#8b5cf6","#ec4899"]
RISK_PAL    = {"CRITICAL":"#dc2626","HIGH":"#ea580c","MEDIUM":"#ca8a04","LOW":"#16a34a"}

# ══════════════════════════════════════════════════════════════════
#  HTML RENDERERS
# ══════════════════════════════════════════════════════════════════

def _th(label, align="left"):
    return (f'<th style="padding:9px 12px;font-size:9px;font-weight:700;text-transform:uppercase;'
            f'letter-spacing:.6px;color:#64748b;text-align:{align};white-space:nowrap;'
            f'background:#0f1f38;border-bottom:2px solid #1e3050">{_html.escape(label)}</th>')

def _risk_badge(lvl):
    c = RISK_PAL.get(lvl,"#475569")
    return (f'<span style="background:{c};color:#fff;padding:2px 9px;border-radius:99px;'
            f'font-size:9.5px;font-weight:800;letter-spacing:.5px;white-space:nowrap">{lvl}</span>')

def _conf_bar(v):
    c = "#22c55e" if v>=90 else "#f59e0b" if v>=70 else "#ef4444"
    return (f'<div style="display:flex;flex-direction:column;align-items:center;gap:2px">'
            f'<span style="color:{c};font-size:11px;font-weight:700">{v}%</span>'
            f'<div style="width:40px;height:3px;background:#1e293b;border-radius:99px">'
            f'<div style="width:{max(4,v)}%;height:100%;background:{c};border-radius:99px"></div>'
            f'</div></div>')

def _fmt_fines(v) -> str:
    """Normalize fines to '$X,XXX' string regardless of int/str input."""
    if v is None or v == "" or v == "N/A" or v == "—":
        return "—"
    if isinstance(v, (int, float)):
        return f"${int(v):,}" if v > 0 else "—"
    s = str(v).strip()
    if s.startswith("$"):
        return s
    try:
        return f"${int(float(s)):,}"
    except Exception:
        return s or "—"


def _safe_url(url: str) -> str:
    """Return url only if it's http/https — else empty string (blocks javascript: XSS)."""
    if url and url.lower().startswith(("http://", "https://")):
        return url
    return ""


def _src_link(url):
    url = _safe_url(url)
    if not url: return '<span style="color:#334155">—</span>'
    try: host = url.split("/")[2].replace("www.","")
    except: host = "link"
    return (f'<a href="{_html.escape(url)}" target="_blank" rel="noopener" '
            f'style="color:#60a5fa;text-decoration:none;font-size:10.5px;font-weight:500">'
            f'{_html.escape(host[:20])} ↗</a>')


def render_regulations_table(rows: list[dict], page: int, per_page: int) -> str:
    total = len(rows)
    start = (page - 1) * per_page
    chunk = rows[start:start + per_page]

    # Fix 1: split into official and INTEL groups
    official = [r for r in chunk if r.get("source_type") != "TELEGRAM_INTEL"]
    intel    = [r for r in chunk if r.get("source_type") == "TELEGRAM_INTEL"]

    def _build_rows(group: list[dict], offset: int) -> str:
        out = ""
        for i, r in enumerate(group):
            bg  = "#0d1b2e" if (i + offset) % 2 == 0 else "#0c1525"
            ai  = r.get("ai_data", {})
            lvl = r["critical_level"]
            rid = r["id"]
            has_ai = r.get("has_ai", False)

            conflict_icon = (
                '<span title="Internal policy conflict" '
                'style="color:#fbbf24;font-size:13px;margin-left:4px">⚠</span>'
                if r.get("conflict") else ""
            )
            act_cnt = r.get("actions", 0)
            act_cell = (
                f'<span style="background:#1e3a8a;color:#93c5fd;padding:2px 7px;'
                f'border-radius:99px;font-size:9.5px;font-weight:700">{act_cnt} steps</span>'
                if act_cnt else '<span style="color:#334155;font-size:10px">—</span>'
            )

            # Fix 7: urgency label
            urgency = r.get("urgency", "—")
            _urg_days = None
            _m = re.match(r"^(\d+)d", urgency)
            if _m:
                _urg_days = int(_m.group(1))
            urgency_color = (
                "#ef4444" if "Overdue" in urgency or "TODAY" in urgency else
                "#f59e0b" if (_urg_days is not None and _urg_days <= 14) else
                "#475569"
            )

            # Fix 6: current status badge
            st = r.get("status", "ACTIVE")
            st_colors = {"ACTIVE":"#1e3a5f:#93c5fd", "ACKNOWLEDGED":"#1a3a1a:#86efac",
                         "IN_PROGRESS":"#3a2a05:#fcd34d", "DONE":"#1e293b:#475569",
                         "HUMAN_REVIEW":"#3a1a05:#fb923c"}
            st_cfg = st_colors.get(st, st_colors["ACTIVE"]).split(":")
            status_badge = (
                f'<span id="statbadge{rid}" style="background:{st_cfg[0]};color:{st_cfg[1]};'
                f'padding:2px 6px;border-radius:99px;font-size:9px;font-weight:700">'
                f'{st.replace("_"," ")}</span>'
            )

            # Main row
            out += (
                f'<tr onclick="var d=document.getElementById(\'d{rid}\');'
                f'd.style.display=d.style.display===\'table-row\'?\'none\':\'table-row\'"'
                f' style="background:{bg};border-bottom:1px solid #162033;cursor:pointer"'
                f' onmouseover="this.style.background=\'#0f2440\'"'
                f' onmouseout="this.style.background=\'{bg}\'">'
                f'<td style="padding:8px 11px;font-weight:600;font-size:11.5px;color:#94a3b8;white-space:nowrap">'
                f'{_html.escape(r["jurisdiction"])}</td>'
                f'<td style="padding:8px 11px;text-align:center">{_risk_badge(lvl)}</td>'
                f'<td style="padding:8px 11px;font-size:11.5px;color:#e2e8f0;max-width:240px;line-height:1.45">'
                f'{_html.escape(r["title"])}{conflict_icon}</td>'
                f'<td style="padding:8px 11px;text-align:center;font-size:11px;'
                f'font-weight:700;color:#ef4444;white-space:nowrap">'
                f'{_html.escape(str(r["fines"]))}</td>'
                f'<td style="padding:8px 11px;text-align:center">{act_cell}</td>'
                f'<td style="padding:8px 11px;font-size:10px;color:{urgency_color};'
                f'text-align:center;white-space:nowrap;font-weight:600">{urgency}</td>'
                f'<td style="padding:8px 11px;text-align:center">'
                f'{_conf_bar(round(r["confidence"] * 100))}</td>'
                f'<td style="padding:8px 5px;text-align:center">{status_badge}</td>'
                f'</tr>'
            )

            # Fix 2: empty AI → pending notice. Fix 4: source link. Fix 6: status buttons. Fix 7: pub date.
            legal   = _html.escape(ai.get("legal", "")) if has_ai else ""
            risk    = _html.escape(ai.get("risk",  "")) if has_ai else ""
            fines_v = _html.escape(str(ai.get("fines", "—")))
            conflict_txt = _html.escape(ai.get("internal_conflict", ""))

            if has_ai:
                acts_html = "".join(
                    f'<div style="display:flex;gap:7px;align-items:flex-start;padding:3px 0">'
                    f'<span style="background:#1e3a8a;color:#93c5fd;padding:1px 5px;'
                    f'border-radius:3px;font-size:9px;font-weight:700;flex-shrink:0;margin-top:2px">'
                    f'{j + 1}</span>'
                    f'<span style="font-size:11.5px;color:#cbd5e1;line-height:1.5">'
                    f'{_html.escape(act)}</span></div>'
                    for j, act in enumerate(ai.get("action", []))
                )
                left_col = (
                    f'<div style="font-size:9px;font-weight:700;letter-spacing:.6px;color:#3b82f6;'
                    f'text-transform:uppercase;margin-bottom:6px">Legal Analysis</div>'
                    f'<div style="font-size:11.5px;color:#cbd5e1;line-height:1.6;background:#0c1a2e;'
                    f'padding:10px 12px;border-radius:8px;border-left:3px solid #3b82f6">{legal}</div>'
                    f'<div style="font-size:9px;font-weight:700;letter-spacing:.6px;color:#ea580c;'
                    f'text-transform:uppercase;margin:10px 0 6px">Risk & Financial Impact</div>'
                    f'<div style="font-size:11.5px;color:#fed7aa;line-height:1.6;background:#1a0e05;'
                    f'padding:10px 12px;border-radius:8px;border-left:3px solid #ea580c">'
                    f'{risk}<br><br>'
                    f'<span style="font-size:13px;font-weight:700;color:#ef4444">Max Fine: {fines_v}</span>'
                    f'</div>'
                )
                right_col = (
                    f'<div style="font-size:9px;font-weight:700;letter-spacing:.6px;color:#22c55e;'
                    f'text-transform:uppercase;margin-bottom:6px">Action Plan</div>'
                    f'<div style="background:#0a1a0e;padding:10px 12px;border-radius:8px;'
                    f'border-left:3px solid #22c55e">{acts_html}</div>'
                )
            else:
                # Fix 2: pending notice
                left_col = (
                    f'<div style="background:#0c1525;border:1px dashed #1e3050;border-radius:8px;'
                    f'padding:16px;text-align:center;color:#334155">'
                    f'<div style="font-size:20px;margin-bottom:6px">⏳</div>'
                    f'<div style="font-size:11px;font-weight:700;color:#475569">AI Analysis Pending</div>'
                    f'<div style="font-size:10px;color:#334155;margin-top:4px">'
                    f'Run pipeline to generate Legal, Risk &amp; Action plan</div>'
                    f'</div>'
                )
                right_col = ""

            conflict_block = (
                f'<div style="background:#422006;border:1px solid #92400e;border-radius:8px;'
                f'padding:10px 14px;margin-top:10px">'
                f'<div style="color:#fbbf24;font-size:9px;font-weight:700;letter-spacing:.5px;'
                f'margin-bottom:4px">INTERNAL POLICY CONFLICT DETECTED</div>'
                f'<div style="color:#fed7aa;font-size:11.5px;line-height:1.55">{conflict_txt}</div>'
                f'</div>'
            ) if conflict_txt else ""

            # Fix 4: source document link (XSS-safe: only http/https)
            src_url = _safe_url(r.get("source", ""))
            src_btn = (
                f'<a href="{_html.escape(src_url)}" target="_blank" rel="noopener" '
                f'style="display:inline-flex;align-items:center;gap:5px;padding:5px 12px;'
                f'background:#0f1f38;border:1px solid #1e3a5f;border-radius:6px;'
                f'color:#60a5fa;font-size:11px;font-weight:600;text-decoration:none;'
                f'transition:border-color .2s" '
                f'onmouseover="this.style.borderColor=\'#3b82f6\'" '
                f'onmouseout="this.style.borderColor=\'#1e3a5f\'">'
                f'Open source document ↗</a>'
            ) if src_url else ""

            # Fix 6: status workflow buttons (fetch API)
            status_btns = "".join(
                f'<button onclick="event.stopPropagation();'
                f'fetch(\'/api/reg/status?id={rid}&status={sv}\').then(()=>{{'
                f'document.getElementById(\'statbadge{rid}\').innerText=\'{sl}\';'
                f'document.getElementById(\'statbadge{rid}\').style.background=\'{sbg}\';'
                f'document.getElementById(\'statbadge{rid}\').style.color=\'{sfg}\'}})" '
                f'style="padding:4px 10px;background:{sbg};color:{sfg};border:none;'
                f'border-radius:6px;font-size:10px;font-weight:700;cursor:pointer;'
                f'transition:opacity .15s" onmouseover="this.style.opacity=\'.75\'" '
                f'onmouseout="this.style.opacity=\'1\'">{sl}</button>'
                for sv, sl, sbg, sfg in [
                    ("ACKNOWLEDGED", "✓ Acknowledged", "#1a3a1a", "#86efac"),
                    ("IN_PROGRESS",  "↻ In Progress",  "#3a2a05", "#fcd34d"),
                    ("DONE",         "✓✓ Done",         "#1e293b", "#64748b"),
                ]
            )

            # Fix 7: metadata bar
            pub = _html.escape(r.get("pub_date", "—"))
            meta_bar = (
                f'<div style="display:flex;gap:16px;padding:8px 0;border-top:1px solid #1e3050;'
                f'margin-top:12px;align-items:center;flex-wrap:wrap">'
                f'<span style="font-size:10px;color:#475569">📅 Published: <b style="color:#64748b">{pub}</b></span>'
                f'<span style="font-size:10px;color:#475569">⏰ Deadline: '
                f'<b style="color:{urgency_color}">{_html.escape(r.get("urgency","—"))}</b></span>'
                f'<span style="font-size:10px;color:#475569">🌐 Jurisdiction: '
                f'<b style="color:#94a3b8">{_html.escape(r.get("jur_raw","—"))}</b></span>'
                f'{src_btn}'
                f'<div style="margin-left:auto;display:flex;gap:6px">{status_btns}</div>'
                f'</div>'
            )

            out += (
                f'<tr id="d{rid}" style="display:none;background:#070f1d">'
                f'<td colspan="8" style="padding:16px 20px;border-bottom:2px solid #1e3a5f">'
                f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">'
                f'<div>{left_col}</div>'
                f'<div>{right_col}{conflict_block}</div>'
                f'</div>'
                f'{meta_bar}'
                f'</td></tr>'
            )
        return out

    # Build section: official docs
    tbody = _build_rows(official, 0)

    # Fix 1: INTEL separator + rows
    if intel:
        intel_banner = (
            f'<tr style="background:#1a0e00">'
            f'<td colspan="8" style="padding:8px 16px;border-top:2px solid #b45309;'
            f'border-bottom:1px solid #78350f">'
            f'<div style="display:flex;align-items:center;gap:10px">'
            f'<span style="background:#b45309;color:#fef3c7;padding:2px 8px;border-radius:4px;'
            f'font-size:9px;font-weight:800;letter-spacing:.08em">UNVERIFIED INTELLIGENCE</span>'
            f'<span style="font-size:10.5px;color:#92400e">'
            f'Sources below are from informal channels (Telegram, industry contacts). '
            f'Not officially verified. For monitoring only — do not use as basis for compliance decisions.</span>'
            f'</div></td></tr>'
        )
        tbody += intel_banner + _build_rows(intel, len(official))

    if not chunk:
        tbody = ('<tr><td colspan="8" style="padding:40px;text-align:center;'
                 'color:#334155;font-size:13px;font-style:italic">No records match filters</td></tr>')

    footer = (
        '<div style="padding:7px 4px;font-size:10px;color:#334155;text-align:right">'
        'No documents found</div>'
        if total == 0 else
        f'<div style="padding:7px 4px;font-size:10px;color:#334155;text-align:right">'
        f'Showing {start + 1}–{min(start + per_page, total)} of {total}</div>'
    )

    return (
        f'<div style="overflow-x:auto;border-radius:10px;border:1px solid #1e3050;overflow:hidden">'
        f'<table style="width:100%;border-collapse:collapse;font-family:Inter,sans-serif">'
        f'<thead><tr>'
        f'{_th("Region")}{_th("Risk","center")}{_th("Document")}'
        f'{_th("Max Fine","center")}{_th("Actions","center")}'
        f'{_th("Deadline","center")}{_th("AI Conf.","center")}{_th("Status","center")}'
        f'</tr></thead>'
        f'<tbody>{tbody}</tbody>'
        f'</table></div>{footer}'
    )


def render_sessions_table(sessions: list) -> str:
    SC = {"RUNNING":"#ca8a04","DONE":"#16a34a","ERROR":"#dc2626"}
    tbody = ""
    for i, s in enumerate(reversed(sessions)):
        bg = "#0d1b2e" if i%2==0 else "#0c1525"
        sc = SC.get(s.status,"#475569")
        badge = (f'<span style="background:{sc};color:#fff;padding:2px 8px;'
                 f'border-radius:6px;font-size:9.5px;font-weight:700">{s.status}</span>')
        # Per-regulator detail
        runs_html = "".join(
            f'<div style="display:flex;gap:6px;align-items:center;padding:2px 0">'
            f'<span style="font-size:10px;color:#475569;font-family:monospace">{FLAGS.get(r.jur,"")}{r.jur}</span>'
            f'<span style="font-size:10px;color:{"#22c55e" if r.status=="OK" else "#ef4444"}">'
            f'{r.status}</span>'
            f'<span style="font-size:9.5px;color:#334155">'
            f'↳ {r.saved} saved · {r.bytes_dl//1024}kb · {r.method}</span>'
            f'</div>'
            for r in s.runs
        ) or '<span style="color:#334155;font-size:10.5px">No runs</span>'

        tbody += (
            f'<tr style="background:{bg};border-bottom:1px solid #162033">'
            f'<td style="padding:9px 12px;font-family:monospace;font-size:12px;color:#60a5fa;font-weight:700">'
            f'{_html.escape(s.sid)}</td>'
            f'<td style="padding:9px 12px;font-size:10.5px;color:#64748b;white-space:nowrap">'
            f'{s.started.strftime("%d.%m.%Y  %H:%M:%S")}</td>'
            f'<td style="padding:9px 12px;text-align:center">{badge}</td>'
            f'<td style="padding:9px 12px;font-size:11.5px;color:#22c55e;font-weight:700;text-align:center">'
            f'{s.saved}</td>'
            f'<td style="padding:9px 12px;font-size:10.5px;color:#64748b;text-align:center;font-family:monospace">'
            f'{s.bytes_dl//1024} KB</td>'
            f'<td style="padding:9px 12px;font-size:10.5px;color:#475569;text-align:center;font-family:monospace">'
            f'{s.duration}</td>'
            f'<td style="padding:9px 12px">{runs_html}</td>'
            f'</tr>'
        )

    if not tbody:
        tbody = ('<tr><td colspan="7" style="padding:36px;text-align:center;'
                 'color:#334155;font-size:13px;font-style:italic">'
                 'No sessions yet. Click "Run Monitoring" to start.</td></tr>')

    return (
        f'<div style="overflow-x:auto;border-radius:10px;border:1px solid #1e3050;overflow:hidden">'
        f'<table style="width:100%;border-collapse:collapse;font-family:Inter,sans-serif">'
        f'<thead><tr>'
        f'{_th("Session ID")}{_th("Started")}{_th("Status","center")}'
        f'{_th("Saved","center")}{_th("Downloaded","center")}'
        f'{_th("Duration","center")}{_th("Regulators Detail")}'
        f'</tr></thead><tbody>{tbody}</tbody></table></div>'
    )

# ══════════════════════════════════════════════════════════════════
#  CSS
# ══════════════════════════════════════════════════════════════════

CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
body,*{font-family:'Inter','Segoe UI',sans-serif!important}
.rr-log{font-family:'Courier New',monospace!important;font-size:11.5px!important;
  line-height:1.75!important;background:#020617!important;color:#34d399!important;
  border-radius:10px!important;padding:12px 14px!important;border:1px solid #1e3a5f!important}
.rr-kpi{transition:transform .15s,box-shadow .15s;cursor:default}
.rr-kpi:hover{transform:translateY(-3px);box-shadow:0 8px 32px rgba(0,0,0,.5)!important}
.rr-nav{display:flex;align-items:center;gap:12px;padding:10px 12px;border-radius:12px;
  cursor:pointer;transition:background .15s,color .15s;white-space:nowrap}
.rr-nav:hover{background:#1e293b}
.rr-nav.active{background:#1e293b;color:#60a5fa!important}
.rr-sec{font-size:10px;font-weight:700;letter-spacing:.12em;color:#334155;
  text-transform:uppercase;padding:4px 4px 8px}
.rr-input input{background:#0f1f38!important;color:#e2e8f0!important;
  border:1px solid #1e3050!important;border-radius:8px!important}
.rr-input label{color:#64748b!important}
::-webkit-scrollbar{width:4px;height:4px}
::-webkit-scrollbar-thumb{background:#1e3a5f;border-radius:99px}
/* Trust signal pulse dot */
@keyframes rr-pulse{0%,100%{opacity:1}50%{opacity:.4}}
.rr-live{animation:rr-pulse 2s ease-in-out infinite}
/* Source badges */
.rr-src-badge{display:inline-flex;align-items:center;gap:5px;padding:4px 10px;
  border-radius:6px;background:#0f1f38;border:1px solid #1e3050;
  font-size:10px;font-weight:700;color:#64748b;letter-spacing:.03em;white-space:nowrap}
/* CTA button glow */
.rr-cta-primary{box-shadow:0 0 20px rgba(37,99,235,.45)!important;
  transition:box-shadow .2s,transform .15s!important}
.rr-cta-primary:hover{box-shadow:0 0 32px rgba(37,99,235,.7)!important;transform:translateY(-1px)}
.rr-cta-outline{border:1px solid #1e3a5f!important;transition:border-color .2s,background .2s!important}
.rr-cta-outline:hover{border-color:#2563eb!important;background:#0f1f38!important}
/* Alert banner */
.rr-alert-banner{background:linear-gradient(90deg,#450a0a,#7f1d1d);
  border:1px solid #dc2626;border-radius:12px;padding:12px 16px}
/* Responsive: hide sidebar on small screens */
@media(max-width:768px){
  .q-drawer{display:none!important}
  .rr-hamburger{display:flex!important}
}
.rr-hamburger{display:none;cursor:pointer;padding:6px;border-radius:8px;
  background:#1e293b;border:1px solid #334155}
/* Stat row in header */
.rr-hdr-stat{display:flex;align-items:center;gap:5px;padding:4px 10px;
  border-radius:6px;background:#0f1820;border:1px solid #1e3050}
</style>
"""

# ══════════════════════════════════════════════════════════════════
#  APP
# ══════════════════════════════════════════════════════════════════

class RegRadarApp:
    def __init__(self):
        init_app_db()
        self._page         = "dashboard"
        self._panels: dict = {}
        self._nav_items: dict = {}
        self._risk_lbls: dict = {}
        self._reg_cbs: dict   = {}
        self._table_page      = 1
        self._per_page        = 10
        self._all_rows: list  = []
        self._sessions: list  = []   # list[ScrapeSession]
        self._drawer          = None  # left drawer ref for mobile toggle
        self._hdr_sync_lbl    = None  # header last-sync label
        self._filter_jur      = "ALL"
        self._filter_lvl      = "ALL"
        self._filter_src      = "ALL"  # ALL / OFFICIAL / INTEL

    # ══ DATA ══════════════════════════════════════════════════════

    def _metrics(self) -> dict:
        with SessionLocal() as db:
            total = db.query(func.count(RegulationRecord.id)).scalar() or 0
            crit  = db.query(func.count(RegulationRecord.id)).filter(
                RegulationRecord.critical_level=="CRITICAL").scalar() or 0
            high  = db.query(func.count(RegulationRecord.id)).filter(
                RegulationRecord.critical_level=="HIGH").scalar() or 0
            jurs  = db.query(RegulationRecord.jurisdiction).distinct().count() or 0

        # Business Value: суммируем штрафы по активным документам
        total_exposure = 0
        with engine.connect() as c:
            rows = c.execute(text(
                "SELECT COALESCE(ai_analysis,'{}') ai_analysis "
                "FROM regulations WHERE status NOT IN ('DONE')"
            )).fetchall()
        for r in rows:
            try:
                ai = json.loads(r.ai_analysis or "{}")
                fv = ai.get("fines")
                if fv and str(fv) not in ("", "N/A", "—"):
                    raw = str(fv).replace("$","").replace(",","").strip()
                    total_exposure += int(float(raw))
            except Exception:
                pass

        # Hours Saved: каждый обработанный документ = 1.5 ч ручной работы аналитика
        hours_saved = total * 1.5

        return {
            "total":          total,
            "critical":       crit,
            "high":           high,
            "jurs":           jurs,
            "risk_exposure":  total_exposure,
            "hours_saved":    hours_saved,
        }

    def _grid_data(self) -> list[dict]:
        with engine.connect() as c:
            rows = c.execute(text(
                "SELECT id,title,jurisdiction,critical_level,effective_date,publication_date,"
                "confidence,source_url,status,"
                "COALESCE(ai_analysis,'{}') ai_analysis,"
                "COALESCE(source_type,'WEB') source_type "
                "FROM regulations ORDER BY "
                "CASE critical_level WHEN 'CRITICAL' THEN 0 WHEN 'HIGH' THEN 1 "
                "WHEN 'MEDIUM' THEN 2 ELSE 3 END, effective_date ASC"
            )).fetchall()
        today = datetime.utcnow().date()
        out = []
        for r in rows:
            j = r.jurisdiction or ""
            try: ai = json.loads(r.ai_analysis or "{}")
            except: ai = {}
            # Compute days-until with deadline label (fix 7)
            eff_date = r.effective_date or ""
            urgency_str = "—"
            try:
                d = datetime.strptime(eff_date, "%Y-%m-%d").date()
                delta = (d - today).days
                if delta < 0:
                    urgency_str = f"Overdue {abs(delta)}d"
                elif delta == 0:
                    urgency_str = "TODAY"
                else:
                    urgency_str = f"{delta}d until {d.strftime('%d %b %Y')}"
            except Exception:
                pass
            out.append({
                "id":             r.id,
                "jurisdiction":   f"{FLAGS.get(j,'')}{j}".strip(),
                "jur_raw":        j,
                "critical_level": (r.critical_level or "LOW").upper(),
                "title":          r.title or "",
                "fines":          _fmt_fines(ai.get("fines")),
                "actions":        len(ai.get("action",[])),
                "conflict":       bool(ai.get("internal_conflict","").strip()),
                "date":           eff_date,
                "urgency":        urgency_str,
                "pub_date":       r.publication_date or "—",
                "confidence":     float(r.confidence or 0),
                "source":         r.source_url or "",
                "source_type":    r.source_type or "WEB",
                "status":         r.status or "ACTIVE",
                "ai_data":        ai,
                "has_ai":         bool(ai.get("legal") or ai.get("action")),
            })
        return out

    def _chart_data(self) -> list[dict]:
        with SessionLocal() as db:
            rows = db.query(RegulationRecord.jurisdiction,
                            func.count(RegulationRecord.id)).group_by(
                RegulationRecord.jurisdiction).all()
        return [{"value":cnt,"name":f"{FLAGS.get(j,'')}{j}".strip()} for j,cnt in rows]

    def _risk_counts(self) -> list[tuple]:
        with SessionLocal() as db:
            return [(lvl, RISK_PAL[lvl],
                     db.query(func.count(RegulationRecord.id)).filter(
                         RegulationRecord.critical_level==lvl).scalar() or 0)
                    for lvl in ("CRITICAL","HIGH","MEDIUM","LOW")]

    def _heatmap_data(self) -> dict:
        levels = ["CRITICAL","HIGH","MEDIUM","LOW"]
        with SessionLocal() as db:
            rows = db.query(RegulationRecord.jurisdiction,
                            RegulationRecord.critical_level,
                            func.count(RegulationRecord.id)).group_by(
                RegulationRecord.jurisdiction,
                RegulationRecord.critical_level).all()
        jurs = sorted({r[0] for r in rows if r[0]})
        cell_map: dict = {}
        for j,lv,cnt in rows:
            cell_map[(j,lv)] = cnt
        data = []
        max_v = 0
        for ji, jur in enumerate(jurs):
            for li, lvl in enumerate(levels):
                v = cell_map.get((jur,lvl), 0)
                data.append([ji, li, v])
                max_v = max(max_v, v)
        return {"jurs":jurs,"levels":levels,"data":data,"max":max_v}

    def _trend_data(self) -> dict:
        with engine.connect() as c:
            rows = c.execute(text(
                "SELECT strftime('%Y-%m',extracted_at) mo,jurisdiction,COUNT(*) cnt "
                "FROM regulations GROUP BY mo,jurisdiction ORDER BY mo"
            )).fetchall()
        months = sorted({r[0] for r in rows if r[0]})
        jurs   = sorted({r[1] for r in rows if r[1]})
        mmap: dict = {}
        for r in rows:
            mmap.setdefault(r[0],{})[r[1]] = r[2]
        series = []
        for idx, jur in enumerate(jurs):
            series.append({
                "name": f"{FLAGS.get(jur,'')}{jur}",
                "type": "line", "smooth": True,
                "data": [mmap.get(m,{}).get(jur,0) for m in months],
                "itemStyle": {"color": PIE_COLORS[idx%len(PIE_COLORS)]},
                "areaStyle": {"opacity": 0.08},
            })
        return {"months": months, "series": series}

    # ══ TABLE PAGINATION ══════════════════════════════════════════

    def _refresh_table(self):
        self._all_rows = self._grid_data()
        rows = self._all_rows
        # Apply filters (fix 5)
        if self._filter_jur != "ALL":
            rows = [r for r in rows if r["jur_raw"] == self._filter_jur]
        if self._filter_lvl != "ALL":
            rows = [r for r in rows if r["critical_level"] == self._filter_lvl]
        if self._filter_src == "OFFICIAL":
            rows = [r for r in rows if r["source_type"] != "TELEGRAM_INTEL"]
        elif self._filter_src == "INTEL":
            rows = [r for r in rows if r["source_type"] == "TELEGRAM_INTEL"]
        tp = math.ceil(len(rows)/self._per_page) or 1
        if self._table_page > tp:
            self._table_page = tp
        self._page_lbl.set_text(
            f"Page {self._table_page}/{tp}  ({len(rows)} docs)"
        )
        self._tbl_wrap.clear()
        with self._tbl_wrap:
            ui.html(render_regulations_table(rows, self._table_page, self._per_page))

    def _prev_page(self):
        if self._table_page > 1:
            self._table_page -= 1; self._refresh_table()

    def _next_page(self):
        tp = math.ceil(len(self._all_rows)/self._per_page) or 1
        if self._table_page < tp:
            self._table_page += 1; self._refresh_table()

    def _refresh_sessions(self):
        self._sess_wrap.clear()
        with self._sess_wrap:
            ui.html(render_sessions_table(self._sessions))

    # ══ NAVIGATION ════════════════════════════════════════════════

    def _show(self, name: str):
        self._page = name
        for n,p in self._panels.items(): p.set_visibility(n==name)
        for n,item in self._nav_items.items():
            if n==name: item.classes(add="active")
            else: item.classes(remove="active")

    # ══ FULL REFRESH ══════════════════════════════════════════════

    def _refresh_all(self):
        m = self._metrics()
        self._lbl_total.set_text(str(m["total"]))
        self._lbl_crit.set_text(str(m["critical"]))
        self._lbl_high.set_text(str(m["high"]))
        self._lbl_jurs.set_text(str(m["jurs"]))
        self._lbl_exposure.set_text(f"${m['risk_exposure']:,.0f}" if m["risk_exposure"] else "$0")
        self._lbl_hours.set_text(f"{m['hours_saved']:.0f}h")
        self._lbl_sync.set_text(datetime.now().strftime("%d.%m  %H:%M"))
        self._refresh_table()
        self._pie.options["series"][0]["data"] = self._chart_data()
        self._pie.update()
        for lvl,_,cnt in self._risk_counts():
            if lvl in self._risk_lbls:
                self._risk_lbls[lvl].set_text(str(cnt))
        self._refresh_sessions()

    # ══ SCRAPING PIPELINE ═════════════════════════════════════════

    async def _run_sync(self):
        self._btn_sync.disable()
        session = ScrapeSession()
        self._sessions.append(session)
        self._refresh_sessions()

        self._log_wrap.clear()
        with self._log_wrap:
            live = ui.log(max_lines=35).classes("rr-log w-full").style("height:320px")
        self._dlg.open()

        with SessionLocal() as db:
            before_total = db.query(func.count(RegulationRecord.id)).scalar() or 0
            before_crit  = db.query(func.count(RegulationRecord.id)).filter(
                RegulationRecord.critical_level=="CRITICAL").scalar() or 0

        loop  = asyncio.get_running_loop()
        log_q: asyncio.Queue = asyncio.Queue()

        def do_scrape():
            def send(m): loop.call_soon_threadsafe(log_q.put_nowait, m)

            regs = reg_repo_r.all_active()
            if not regs:
                send("ERROR: No active regulators."); send("__DONE__"); return

            send(f"RegRadar Enterprise Pipeline — {len(regs)} regulators")
            send(f"curl_cffi: {'enabled' if _HAS_CURL_CFFI else 'disabled'}  |  "
                 f"AI multi-agent: {'enabled' if os.getenv('OPENAI_API_KEY') else 'disabled'}")
            send("")

            for reg in regs:
                run = RegRun(name=reg.name, jur=reg.jurisdiction)
                session.runs.append(run)
                send(f"[{reg.jurisdiction}] ▶  {reg.name}")
                send(f"[{reg.jurisdiction}] GET  {reg.base_url}")
                try:
                    saved = _real_scrape(
                        reg.name, reg.base_url, reg.jurisdiction, reg.link_pattern
                    )
                    run.saved  = saved
                    run.found  = saved
                    run.status = "OK"
                    run.method = "requests"
                    send(f"[{reg.jurisdiction}] OK   {saved} new docs saved")
                except Exception as e:
                    err = str(e)
                    if _HAS_CURL_CFFI and any(x in err for x in
                            ["403","404","Forbidden","blocked","SSL","timeout"]):
                        send(f"[{reg.jurisdiction}] WARN {err[:60]}")
                        send(f"[{reg.jurisdiction}] ↩   Retrying via curl-cffi...")
                        try:
                            saved = _fallback_scrape(
                                reg.base_url, reg.jurisdiction, reg.link_pattern
                            )
                            run.saved  = saved
                            run.status = "OK"
                            run.method = "curl_cffi"
                            send(f"[{reg.jurisdiction}] OK   curl-cffi: {saved} docs")
                        except Exception as e2:
                            run.status = "FAIL"
                            run.error  = str(e2)[:80]
                            send(f"[{reg.jurisdiction}] FAIL {run.error}")
                    else:
                        run.status = "FAIL"
                        run.error  = err[:80]
                        send(f"[{reg.jurisdiction}] FAIL {run.error}")
                send("")

            send("Pipeline complete.")
            send("__DONE__")

        future = loop.run_in_executor(
            concurrent.futures.ThreadPoolExecutor(max_workers=3), do_scrape
        )
        while True:
            try:
                msg = await asyncio.wait_for(log_q.get(), timeout=0.25)
            except TimeoutError:
                if future.done(): break
                continue
            if msg == "__DONE__": break
            try:
                live.push(f"[{datetime.now().strftime('%H:%M:%S')}]  {msg}")
            except RuntimeError:
                pass  # client disconnected mid-scrape; data is still saved to DB

        try:
            await future
        except Exception as e:
            try:
                live.push(f"FATAL: {e}")
            except RuntimeError:
                pass
            session.status = "ERROR"

        session.status   = "ERROR" if any(r.status=="FAIL" for r in session.runs) else "DONE"
        session.finished = datetime.now()

        await asyncio.sleep(0.4)
        try:
            self._dlg.close()
            self._refresh_all()
        except RuntimeError:
            pass

        with SessionLocal() as db:
            after_total = db.query(func.count(RegulationRecord.id)).scalar() or 0
            after_crit  = db.query(func.count(RegulationRecord.id)).filter(
                RegulationRecord.critical_level=="CRITICAL").scalar() or 0

        added    = after_total - before_total
        new_crit = after_crit  - before_crit

        if new_crit > 0:
            new_crit_rows = self._grid_data()[:new_crit]
            await _dispatch_alerts(new_crit_rows)

        try:
            ui.notify(
                f"Done. Added {added} docs, {new_crit} new CRITICAL threats.",
                type="negative" if new_crit > 0 else "positive",
                position="top-right", timeout=7000,
            )
            self._btn_sync.enable()
        except RuntimeError:
            pass

    # ══ EXCEL EXPORT ══════════════════════════════════════════════

    def _export_excel(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rows = self._grid_data()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RegRadar Export"

        HDRS  = ["#","Region","Risk","Document Title","Legal Analysis",
                 "Risk & Fines","Max Fine","Action Plan","Conflict","Date","Conf.","Source"]
        WIDTHS= [5,10,14,48,55,45,14,55,40,14,10,35]
        THIN  = Side(style="thin", color="D1D5DB")
        BORD  = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
        H_FILL= PatternFill("solid",fgColor="0F2040")
        H_FONT= Font(bold=True,color="FFFFFF",name="Calibri",size=10)
        ROW_F = {"CRITICAL":PatternFill("solid",fgColor="FFD0D0"),
                 "HIGH":    PatternFill("solid",fgColor="FFE5CC"),
                 "MEDIUM":  PatternFill("solid",fgColor="FFFACC")}
        RCOL  = {"CRITICAL":"B91C1C","HIGH":"C05621","MEDIUM":"92400E","LOW":"065F46"}

        for col,(h,w) in enumerate(zip(HDRS,WIDTHS),1):
            c = ws.cell(1,col,h)
            c.font=H_FONT; c.fill=H_FILL; c.border=BORD
            c.alignment=Alignment(horizontal="center",vertical="center")
            ws.column_dimensions[get_column_letter(col)].width=w
        ws.row_dimensions[1].height=26
        ws.freeze_panes="A2"
        ws.auto_filter.ref=f"A1:{get_column_letter(len(HDRS))}1"

        for i,r in enumerate(rows,2):
            lvl = r["critical_level"]
            ai  = r.get("ai_data",{})
            act_text = "\n".join(f"{j+1}. {a}" for j,a in enumerate(ai.get("action",[])))
            summary_len = len(ai.get("legal","")) + len(act_text)
            ws.row_dimensions[i].height = max(16, math.ceil(summary_len/80)*13+4)
            vals = [
                i-1, r["jurisdiction"], lvl, r["title"],
                ai.get("legal",""), ai.get("risk",""),
                r["fines"], act_text,
                ai.get("internal_conflict",""),
                r["date"], f'{round(r["confidence"]*100)}%', r["source"],
            ]
            for col,v in enumerate(vals,1):
                cell = ws.cell(i,col,v)
                cell.border=BORD
                cell.font=Font(name="Calibri",size=9)
                cell.alignment=Alignment(vertical="top",wrap_text=(col in {4,5,6,8,9}),
                                         horizontal="center" if col in {1,3,7,10,11} else "left")
                if ROW_F.get(lvl): cell.fill=ROW_F[lvl]
            ws.cell(i,3).font=Font(bold=True,name="Calibri",size=9,
                                   color=RCOL.get(lvl,"475569"))
            ws.cell(i,3).alignment=Alignment(horizontal="center",vertical="center")

        out = BASE_DIR/"data"/f"regrada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(str(out))
        ui.download(str(out))

    # ══ SETTINGS HELPERS ══════════════════════════════════════════

    def _save_api_key(self, key: str):
        if not key.strip():
            ui.notify("Empty key.", type="warning"); return
        try:
            lines = ENV_PATH.read_text(encoding="utf-8").splitlines() if ENV_PATH.exists() else []
            lines = [l for l in lines if not l.startswith("OPENAI_API_KEY=")]
            lines.append(f"OPENAI_API_KEY={key.strip()}")
            ENV_PATH.write_text("\n".join(lines)+"\n", encoding="utf-8")
            os.environ["OPENAI_API_KEY"] = key.strip()
            ui.notify("API key saved.", type="positive")
        except Exception as e:
            ui.notify(f"Save failed: {e}", type="negative")

    def _save_regulators(self):
        try:
            with SessionLocal() as db:
                for rid,cb in self._reg_cbs.items():
                    r = db.get(Regulator, rid)
                    if r: r.active = cb.value
                db.commit()
            active = sum(1 for cb in self._reg_cbs.values() if cb.value)
            ui.notify(f"Saved. {active} regulators active.", type="positive")
        except Exception as e:
            ui.notify(f"Error: {e}", type="negative")

    def _save_webhooks(self, fields: dict):
        for k,inp in fields.items():
            _wh_set(k, inp.value.strip())
        ui.notify("Integration settings saved.", type="positive")

    async def _test_telegram(self, token_inp, chat_inp):
        ok = await _tg_send(token_inp.value.strip(), chat_inp.value.strip(),
                            "✅ <b>RegRadar Enterprise</b>\n\nTelegram integration test successful.")
        ui.notify("Telegram OK ✓" if ok else "Telegram FAIL — check token/chat_id",
                  type="positive" if ok else "negative")

    def _save_policy(self, name_inp, content_inp, policies_wrap):
        name    = name_inp.value.strip()
        content = content_inp.value.strip()
        if not name or not content:
            ui.notify("Name and content required.", type="warning"); return
        with engine.connect() as c:
            c.execute(text("INSERT INTO company_policies(name,content) VALUES(:n,:c)"),
                      {"n":name,"c":content})
            c.commit()
        name_inp.set_value("")
        content_inp.set_value("")
        self._reload_policies(policies_wrap)
        ui.notify(f"Policy '{name}' saved.", type="positive")

    def _reload_policies(self, wrap):
        wrap.clear()
        with engine.connect() as c:
            rows = c.execute(text("SELECT id,name,created_at FROM company_policies ORDER BY id DESC")).fetchall()
        with wrap:
            if not rows:
                ui.label("No policies uploaded yet.").classes("text-xs text-slate-600")
            for r in rows:
                with ui.row().classes("w-full items-center justify-between py-1 border-b border-slate-800"):
                    with ui.column().classes("gap-0"):
                        ui.label(r[1]).classes("text-sm font-semibold text-white")
                        ui.label(str(r[2])[:16]).classes("text-xs text-slate-600")
                    def _del(rid=r[0], w=wrap):
                        with engine.connect() as cc:
                            cc.execute(text("DELETE FROM company_policies WHERE id=:i"),{"i":rid})
                            cc.commit()
                        self._reload_policies(w)
                        ui.notify("Policy deleted.", type="positive")
                    ui.button("✕", on_click=_del).props("flat round dense size=xs").classes("text-red-500")

    # ══ BUILD UI ══════════════════════════════════════════════════

    def build_ui(self):
        ui.add_head_html(CSS, shared=True)
        ui.colors(primary="#2563eb", secondary="#10b981", negative="#dc2626")
        dm = ui.dark_mode(value=True)

        # ── Pipeline dialog ────────────────────────────────────────
        with ui.dialog().props("persistent") as self._dlg:
            with ui.card().classes("bg-slate-900 border border-slate-700 rounded-2xl p-6").style(
                    "min-width:580px;box-shadow:0 30px 60px rgba(0,0,0,.85)"):
                with ui.row().classes("items-center gap-3 w-full"):
                    ui.spinner("ball", size="lg", color="blue")
                    with ui.column().classes("gap-0"):
                        ui.label("AI Pipeline Running").classes("text-white text-base font-bold")
                        ui.label("Scraper → pdfplumber → GPT-4o-mini → Legal+Risk+Action").classes(
                            "text-slate-500 text-xs").style("font-family:monospace")
                ui.separator().classes("border-slate-700 my-3")
                self._log_wrap = ui.column().classes("w-full")
                with self._log_wrap:
                    ui.log(max_lines=35).classes("rr-log w-full").style("height:320px")
                ui.label("Pipeline running — do not close tab").classes(
                    "text-slate-600 text-xs text-center mt-2 w-full")

        # ── Request Report dialog ──────────────────────────────────
        with ui.dialog() as self._report_dlg:
            with ui.card().classes(
                    "bg-slate-900 border border-slate-700 rounded-2xl p-6 gap-4"
            ).style("min-width:460px"):
                with ui.row().classes("items-center gap-3 w-full"):
                    ui.label("📄").classes("text-2xl")
                    ui.label("Request Regulatory Report").classes("text-white text-base font-bold")
                ui.separator().classes("border-slate-800")
                ui.label("Our analysts will prepare a custom report on selected jurisdictions "
                         "within 24 hours.").classes("text-slate-400 text-sm")
                name_f  = ui.input("Full name / Company").classes("w-full rr-input")
                email_f = ui.input("Work email").classes("w-full rr-input")
                scope_f = ui.select(
                    ["RU — Russia", "KZ — Kazakhstan", "AZ — Azerbaijan",
                     "BY — Belarus", "UZ — Uzbekistan", "All CIS"],
                    label="Jurisdiction scope", value="All CIS",
                ).classes("w-full")
                with ui.row().classes("w-full justify-end gap-2 mt-2"):
                    ui.button("Cancel", on_click=self._report_dlg.close).props(
                        "flat no-caps").classes("text-slate-400")
                    ui.button("Send Request", on_click=lambda: (
                        ui.notify(
                            f"Request sent — we will contact {email_f.value or 'you'} within 24h",
                            type="positive", position="top-right", timeout=5000,
                        ),
                        self._report_dlg.close(),
                    )).props("color=primary unelevated no-caps").classes("font-bold rr-cta-primary")

        # ── Contact Analyst dialog ─────────────────────────────────
        with ui.dialog() as self._analyst_dlg:
            with ui.card().classes(
                    "bg-slate-900 border border-slate-700 rounded-2xl p-6 gap-4"
            ).style("min-width:420px"):
                with ui.row().classes("items-center gap-3 w-full"):
                    ui.label("💬").classes("text-2xl")
                    ui.label("Contact Regulatory Analyst").classes("text-white text-base font-bold")
                ui.separator().classes("border-slate-800")
                ui.html("""
                <div style="font-size:13px;color:#94a3b8;line-height:1.7">
                  Our CIS regulatory experts are available <b style="color:#e2e8f0">Mon–Fri 9:00–18:00 MSK</b>.<br>
                  Average response time: <b style="color:#22c55e">under 2 hours</b>.
                </div>
                <div style="margin-top:14px;display:flex;flex-direction:column;gap:8px">
                  <a href="mailto:compliance@regrada.io" style="color:#60a5fa;font-size:13px;text-decoration:none">
                    ✉ compliance@regrada.io</a>
                  <a href="https://t.me/regrada_support" style="color:#60a5fa;font-size:13px;text-decoration:none">
                    💬 t.me/regrada_support</a>
                </div>
                """)
                with ui.row().classes("w-full justify-end mt-4"):
                    ui.button("Close", on_click=self._analyst_dlg.close).props(
                        "flat no-caps").classes("text-slate-400")

        # ── Header ────────────────────────────────────────────────
        m = self._metrics()
        with ui.header().classes("bg-slate-900 border-b border-slate-800 px-4").style(
                "height:56px;box-shadow:0 2px 20px rgba(0,0,0,.6)"):
            with ui.row().classes("items-center justify-between w-full h-full gap-3"):

                # Left: logo + hamburger (mobile)
                with ui.row().classes("items-center gap-2"):
                    with ui.element("div").classes("rr-hamburger").on(
                            "click", lambda: self._drawer.toggle()):
                        ui.label("☰").style("color:#94a3b8;font-size:18px;line-height:1")
                    ui.label("📡").classes("text-2xl leading-none")
                    ui.label("RegRadar").classes("text-white text-xl font-black tracking-tight")
                    ui.badge("Enterprise", color="blue-grey").props("rounded").classes(
                        "text-xs font-bold hidden md:flex")

                # Center: trust signal stats (hidden on mobile)
                with ui.row().classes("items-center gap-2 hidden md:flex"):
                    with ui.element("div").classes("rr-hdr-stat"):
                        ui.element("div").style(
                            "width:6px;height:6px;border-radius:50%;background:#22c55e;flex-shrink:0"
                        ).classes("rr-live")
                        ui.label("Live").style("font-size:10px;font-weight:700;color:#22c55e;letter-spacing:.04em")
                    with ui.element("div").classes("rr-hdr-stat"):
                        ui.label(f"🏛 {m['jurs']} jurisdictions").style(
                            "font-size:11px;color:#64748b;font-weight:600")
                    with ui.element("div").classes("rr-hdr-stat"):
                        ui.label(f"📄 {m['total']} documents").style(
                            "font-size:11px;color:#64748b;font-weight:600")
                    with ui.element("div").classes("rr-hdr-stat"):
                        ui.label("🕐").style("font-size:11px;color:#475569")
                        self._hdr_sync_lbl = ui.label(
                            "Updated " + datetime.now().strftime("%d %b, %H:%M")
                        ).style("font-size:11px;color:#475569;font-weight:500")

                # Right: dark mode toggle
                ui.button("🌙", on_click=dm.toggle).props("flat round dense").classes(
                    "text-slate-400").style("font-size:17px")

        # ── Sidebar ───────────────────────────────────────────────
        NAV = [
            ("dashboard", "📊", "Risk Dashboard"),
            ("analytics", "📈", "Analytics"),
            ("jobs",      "⏳", "Job History"),
            ("settings",  "⚙️", "Settings"),
        ]
        with ui.left_drawer(value=True).classes(
                "bg-slate-900 text-slate-300 p-4"
        ).style("display:flex;flex-direction:column;justify-content:space-between") as self._drawer:

            with ui.column().classes("w-full gap-1"):
                # Logo in drawer (visible on mobile)
                with ui.row().classes("items-center gap-2 px-1 mb-3 md:hidden"):
                    ui.label("📡").classes("text-lg")
                    ui.label("RegRadar Enterprise").classes("text-white font-black text-sm")

                ui.element("div").classes("rr-sec").add_slot("default", "<span>Navigation</span>")
                for pid, emoji, label in NAV:
                    with ui.element("div").classes(
                            "rr-nav" + (" active" if pid == "dashboard" else "")
                    ).on("click", lambda p=pid: self._show(p)) as item:
                        ui.label(emoji).classes("text-base leading-none w-6 text-center")
                        ui.label(label).classes("text-sm font-medium")
                    self._nav_items[pid] = item

                ui.separator().classes("border-slate-800 my-3")
                ui.element("div").classes("rr-sec").add_slot("default", "<span>Actions</span>")

                # Primary CTA — Update Data
                self._btn_sync = ui.button(
                    "↻  Update Data", on_click=self._run_sync,
                ).classes("w-full font-bold rounded-xl rr-cta-primary").props(
                    "color=primary no-caps unelevated"
                ).style("padding:11px 0;font-size:13px")

                # Secondary CTAs
                ui.button(
                    "📄  Request Report",
                    on_click=self._report_dlg.open,
                ).classes("w-full rounded-xl rr-cta-outline mt-2").props(
                    "flat no-caps"
                ).style("padding:9px 0;font-size:12px;color:#93c5fd;font-weight:600")

                ui.button(
                    "💬  Contact Analyst",
                    on_click=self._analyst_dlg.open,
                ).classes("w-full rounded-xl rr-cta-outline mt-1").props(
                    "flat no-caps"
                ).style("padding:9px 0;font-size:12px;color:#86efac;font-weight:600")

            # Bottom status block
            with ui.column().classes("w-full gap-1 mt-4"):
                ui.separator().classes("border-slate-800 mb-2")

                # Source logos strip
                ui.element("div").classes("rr-sec").add_slot("default", "<span>Data Sources</span>")
                sources = [
                    ("🇷🇺", "ЦБ РФ"), ("🇰🇿", "АРРФР"), ("🇦🇿", "ЦБА"),
                    ("🇧🇾", "НБРБ"), ("🇺🇿", "ЦБ УЗ"),
                ]
                with ui.element("div").style(
                        "display:flex;flex-wrap:wrap;gap:4px;padding:0 2px"):
                    for flag, name in sources:
                        ui.element("div").classes("rr-src-badge").add_slot(
                            "default", f"<span>{flag}</span><span>{name}</span>"
                        )

                ui.separator().classes("border-slate-800 my-2")
                with ui.row().classes("items-center gap-2 px-1"):
                    ui.element("div").style(
                        "width:7px;height:7px;border-radius:50%;background:#22c55e;flex-shrink:0"
                    ).classes("rr-live")
                    ui.label("All systems operational").classes("text-xs text-slate-500")
                with ui.row().classes("items-center gap-2 px-1 mt-1"):
                    ui.label("🕐").classes("text-xs text-slate-600 leading-none")
                    self._lbl_sync = ui.label(
                        datetime.now().strftime("%d.%m  %H:%M")
                    ).classes("text-xs text-slate-600")

        # ── Page container ────────────────────────────────────────
        with ui.column().classes("w-full p-4 md:p-6 bg-slate-950 min-h-screen gap-6"):
            with ui.element("div").classes("w-full") as p1:
                self._panels["dashboard"] = p1; self._build_dashboard()
            with ui.element("div").classes("w-full") as p2:
                self._panels["analytics"] = p2; self._build_analytics()
            with ui.element("div").classes("w-full") as p3:
                self._panels["jobs"] = p3; self._build_jobs()
            with ui.element("div").classes("w-full") as p4:
                self._panels["settings"] = p4; self._build_settings()

        self._show("dashboard")

    # ══ PAGE: DASHBOARD ═══════════════════════════════════════════

    def _build_dashboard(self):
        m = self._metrics()

        # Critical alert banner — only shown when there are CRITICAL threats
        if m["critical"] > 0:
            with ui.element("div").classes("rr-alert-banner w-full"):
                with ui.row().classes("items-center gap-3 w-full"):
                    ui.label("🚨").classes("text-xl flex-shrink-0")
                    with ui.column().classes("gap-0 flex-1"):
                        ui.label(
                            f"{m['critical']} CRITICAL regulation{'s' if m['critical']>1 else ''} "
                            f"require immediate compliance action"
                        ).classes("text-white font-bold text-sm")
                        ui.label(
                            "Review highlighted rows below. Contact your analyst for an urgent briefing."
                        ).classes("text-red-300 text-xs mt-0.5")
                    ui.button(
                        "Contact Analyst →",
                        on_click=self._analyst_dlg.open,
                    ).props("flat no-caps").style(
                        "color:#fca5a5;font-size:12px;font-weight:700;white-space:nowrap")

        with ui.row().classes("w-full gap-4 items-stretch mt-2" if m["critical"] > 0 else "w-full gap-4 items-stretch"):
            # Форматируем Business Value значения
            exposure_str = f"${m['risk_exposure']:,.0f}" if m["risk_exposure"] else "$0"
            hours_str    = f"{m['hours_saved']:.0f}h"
            kpis = [
                ("TOTAL DOCUMENTS","📁",str(m["total"]),
                 "linear-gradient(135deg,#1e3a8a,#2563eb)","border-blue-950/40"),
                ("CRITICAL THREATS","🚨",str(m["critical"]),
                 "linear-gradient(135deg,#7f1d1d,#dc2626)","border-red-950/40"),
                ("HIGH RISK","⚠️",str(m["high"]),
                 "linear-gradient(135deg,#7c2d12,#ea580c)","border-orange-950/40"),
                ("JURISDICTIONS","🌐",str(m["jurs"]),
                 "linear-gradient(135deg,#14532d,#16a34a)","border-green-950/40"),
                ("RISK EXPOSURE","💰",exposure_str,
                 "linear-gradient(135deg,#4a1d96,#7c3aed)","border-violet-950/40"),
                ("HOURS SAVED BY AI","⚡",hours_str,
                 "linear-gradient(135deg,#0c4a6e,#0284c7)","border-sky-950/40"),
            ]
            lbls = []
            for title,emoji,val,grad,border in kpis:
                with ui.card().classes(
                        f"rr-kpi p-4 shadow-md flex-1 rounded-xl border {border}"
                ).style(f"background:{grad}"):
                    with ui.row().classes("w-full justify-between items-center no-wrap"):
                        ui.label(title).classes("text-[10px] font-bold tracking-wider text-white/70 uppercase")
                        ui.label(emoji).classes("text-lg leading-none")
                    lbls.append(ui.label(val).classes("text-3xl font-black text-white mt-2"))
            (self._lbl_total, self._lbl_crit, self._lbl_high,
             self._lbl_jurs, self._lbl_exposure, self._lbl_hours) = lbls

        with ui.row().classes("w-full gap-5 items-start mt-2"):
            # Table card
            with ui.card().classes("flex-1 p-4 rounded-xl shadow-xl bg-slate-900 border border-slate-800").style("min-width:0"):
                with ui.row().classes("w-full justify-between items-center mb-3 no-wrap"):
                    with ui.column().classes("gap-0.5"):
                        with ui.row().classes("items-center gap-2"):
                            ui.label("📋").classes("text-base leading-none")
                            ui.label("Regulatory Changes Feed").classes("text-base font-bold text-white")
                        with ui.row().classes("items-center gap-2 mt-0.5"):
                            ui.element("div").style(
                                "width:5px;height:5px;border-radius:50%;background:#22c55e;flex-shrink:0"
                            ).classes("rr-live")
                            ui.label(
                                f"AI-extracted · {datetime.now().strftime('%d %b %Y, %H:%M')} · "
                                f"Official sources: ЦБ РФ, АРРФР, ЦБА, НБРБ, ЦБ УЗ"
                            ).classes("text-xs text-slate-600")
                    with ui.row().classes("items-center gap-1 flex-shrink-0"):
                        ui.button("◀", on_click=self._prev_page).props("flat round dense").classes("text-slate-400 text-xs")
                        self._page_lbl = ui.label("").classes("text-xs text-slate-500 font-mono px-1")
                        ui.button("▶", on_click=self._next_page).props("flat round dense").classes("text-slate-400 text-xs")
                        ui.button("⬇ Export", on_click=self._export_excel).props(
                            "color=green-6 no-caps unelevated").classes("rounded-lg text-xs font-bold ml-2")
                # ── Filter bar (fix 5) ────────────────────────────
                with ui.row().classes("w-full items-center gap-2 flex-wrap mb-3"):
                    ui.label("Filter:").classes("text-xs text-slate-600 font-bold uppercase tracking-wider flex-shrink-0")

                    # Jurisdiction chips
                    jur_btns = {}
                    def _set_jur(j, btns=None):
                        self._filter_jur = j
                        self._table_page = 1
                        if btns:
                            for k, b in btns.items():
                                b.props("outline" if k != j else "unelevated")
                        self._refresh_table()
                    for jv in ["ALL","RU","KZ","AZ","BY","UZ"]:
                        lbl = jv if jv != "ALL" else "All regions"
                        b = ui.button(lbl, on_click=lambda j=jv: _set_jur(j, jur_btns)
                            ).props("unelevated" if jv == "ALL" else "outline").classes(
                            "text-xs font-bold rounded-lg"
                        ).style("padding:4px 10px;min-height:0;font-size:10.5px")
                        jur_btns[jv] = b

                    ui.separator().props("vertical").classes("border-slate-700 h-5 mx-1")

                    # Risk level chips
                    lvl_colors = {"ALL":"grey-7","CRITICAL":"red-9","HIGH":"orange-9",
                                  "MEDIUM":"yellow-9","LOW":"green-9"}
                    lvl_btns = {}
                    def _set_lvl(lv, btns=None):
                        self._filter_lvl = lv
                        self._table_page = 1
                        if btns:
                            for k, b in btns.items():
                                b.props(f"color={lvl_colors[k]} " + ("unelevated" if k == lv else "outline"))
                        self._refresh_table()
                    for lv in ["ALL","CRITICAL","HIGH","MEDIUM","LOW"]:
                        lbl = lv if lv != "ALL" else "All risks"
                        b = ui.button(lbl, on_click=lambda l=lv: _set_lvl(l, lvl_btns)
                            ).props(f"color={lvl_colors[lv]} " + ("unelevated" if lv == "ALL" else "outline")
                            ).classes("text-xs font-bold rounded-lg"
                            ).style("padding:4px 10px;min-height:0;font-size:10.5px")
                        lvl_btns[lv] = b

                    ui.separator().props("vertical").classes("border-slate-700 h-5 mx-1")

                    # Source type toggle
                    src_btns = {}
                    def _set_src(s, btns=None):
                        self._filter_src = s
                        self._table_page = 1
                        if btns:
                            for k, b in btns.items():
                                b.props("unelevated" if k == s else "outline")
                        self._refresh_table()
                    for sv, sl in [("ALL","All sources"),("OFFICIAL","Official only"),("INTEL","Intel only")]:
                        b = ui.button(sl, on_click=lambda s=sv: _set_src(s, src_btns)
                            ).props("unelevated" if sv == "ALL" else "outline"
                            ).classes("text-xs font-bold rounded-lg"
                            ).style("padding:4px 10px;min-height:0;font-size:10.5px;"
                                    "color:#a78bfa" if sv == "INTEL" else "padding:4px 10px;min-height:0;font-size:10.5px")
                        src_btns[sv] = b

                self._tbl_wrap = ui.element("div").classes("w-full")
                self._refresh_table()

            # Pie + risk legend
            with ui.card().classes("p-4 rounded-xl shadow-xl bg-slate-900 border border-slate-800").style(
                    "min-width:280px;width:300px;flex-shrink:0"):
                with ui.row().classes("items-center gap-2 mb-1"):
                    ui.label("🌐").classes("text-base leading-none")
                    ui.label("Regulatory Coverage").classes("text-base font-bold text-white")
                ui.label("Documents by jurisdiction").classes("text-xs text-slate-600 mb-2")
                self._pie = ui.echart({
                    "backgroundColor":"transparent",
                    "tooltip":{"trigger":"item","formatter":"{b}: {c} ({d}%)",
                               "backgroundColor":"#0f1f38","borderColor":"#1e3a5f",
                               "textStyle":{"color":"#e2e8f0","fontSize":13}},
                    "legend":{"bottom":"2%","left":"center","textStyle":{"color":"#64748b","fontSize":11}},
                    "series":[{"name":"Docs","type":"pie","radius":["42%","70%"],"center":["50%","44%"],
                               "itemStyle":{"borderRadius":8,"borderColor":"#0c1420","borderWidth":3},
                               "label":{"show":False},
                               "emphasis":{"itemStyle":{"shadowBlur":20,"shadowColor":"rgba(59,130,246,.55)"},
                                           "label":{"show":True,"fontSize":14,"fontWeight":"bold","color":"#f1f5f9"}},
                               "data":self._chart_data(),"color":PIE_COLORS}],
                }).classes("w-full h-[280px]")
                ui.separator().classes("border-slate-700 my-2")
                self._risk_lbls = {}
                for lvl,color,cnt in self._risk_counts():
                    with ui.row().classes("w-full items-center justify-between").style("padding:3px 2px"):
                        with ui.row().classes("items-center gap-2"):
                            ui.element("div").style(f"width:8px;height:8px;border-radius:50%;background:{color};flex-shrink:0")
                            ui.label(lvl).style(f"color:{color};font-size:11px;font-weight:700;letter-spacing:.4px")
                        self._risk_lbls[lvl] = ui.label(str(cnt)).classes("text-slate-300 text-sm font-bold")

    # ══ PAGE: ANALYTICS ═══════════════════════════════════════════

    def _build_analytics(self):
        ui.label("Analytics — Risk Intelligence").classes("text-xl font-black text-white mb-4")
        with ui.row().classes("w-full gap-5 items-start"):

            # Left column: heatmap + trend
            with ui.column().classes("flex-1 gap-5").style("min-width:0"):

                # Risk Heatmap
                with ui.card().classes("w-full p-4 rounded-xl bg-slate-900 border border-slate-800"):
                    with ui.row().classes("items-center gap-2 mb-3"):
                        ui.label("🔥").classes("text-base leading-none")
                        ui.label("Risk Heatmap — Jurisdiction × Level").classes("text-base font-bold text-white")
                    hd = self._heatmap_data()
                    ui.echart({
                        "backgroundColor":"transparent",
                        "tooltip":{"position":"top","formatter":"function(p){return p.data[2]+' docs'}",
                                   "backgroundColor":"#0f1f38","textStyle":{"color":"#e2e8f0"}},
                        "grid":{"height":"75%","top":"10%"},
                        "xAxis":{"type":"category","data":hd["jurs"],
                                 "axisLabel":{"color":"#64748b","fontSize":12},"splitArea":{"show":True}},
                        "yAxis":{"type":"category","data":hd["levels"],
                                 "axisLabel":{"color":"#64748b","fontSize":11},"splitArea":{"show":True}},
                        "visualMap":{"min":0,"max":max(hd["max"],1),"calculable":True,
                                     "orient":"horizontal","left":"center","bottom":"5%",
                                     "inRange":{"color":["#0c1525","#1e3a8a","#dc2626"]},
                                     "textStyle":{"color":"#64748b"}},
                        "series":[{"name":"Regulations","type":"heatmap","data":hd["data"],
                                   "label":{"show":True,"color":"#e2e8f0","fontSize":13,"fontWeight":"bold"},
                                   "emphasis":{"itemStyle":{"shadowBlur":10,"shadowColor":"rgba(0,0,0,.5)"}}}],
                    }).classes("w-full h-[260px]")

                # Monthly trend
                with ui.card().classes("w-full p-4 rounded-xl bg-slate-900 border border-slate-800"):
                    with ui.row().classes("items-center gap-2 mb-3"):
                        ui.label("📈").classes("text-base leading-none")
                        ui.label("Regulatory Activity Trend by Month").classes("text-base font-bold text-white")
                    td = self._trend_data()
                    ui.echart({
                        "backgroundColor":"transparent",
                        "tooltip":{"trigger":"axis","backgroundColor":"#0f1f38","borderColor":"#1e3a5f",
                                   "textStyle":{"color":"#e2e8f0"}},
                        "legend":{"textStyle":{"color":"#64748b","fontSize":11},"top":"0"},
                        "grid":{"left":"3%","right":"4%","bottom":"10%","containLabel":True},
                        "xAxis":{"type":"category","data":td["months"],
                                 "axisLabel":{"color":"#64748b","fontSize":11},
                                 "axisLine":{"lineStyle":{"color":"#1e3050"}}},
                        "yAxis":{"type":"value","minInterval":1,"axisLabel":{"color":"#64748b"},
                                 "splitLine":{"lineStyle":{"color":"#1e3050","type":"dashed"}}},
                        "series":td["series"],
                    }).classes("w-full h-[260px]")

            # Right column: stacked bar + risk totals
            with ui.column().classes("gap-5").style("width:280px;flex-shrink:0"):
                with ui.card().classes("w-full p-4 rounded-xl bg-slate-900 border border-slate-800"):
                    with ui.row().classes("items-center gap-2 mb-3"):
                        ui.label("🎯").classes("text-base leading-none")
                        ui.label("Total by Risk Level").classes("text-sm font-bold text-white")
                    rcs = self._risk_counts()
                    mx = max((c for _,_,c in rcs), default=1) or 1
                    for lvl,color,cnt in rcs:
                        with ui.column().classes("w-full gap-1 mb-3"):
                            with ui.row().classes("w-full justify-between items-center"):
                                ui.label(lvl).style(f"color:{color};font-size:11px;font-weight:700")
                                ui.label(str(cnt)).classes("text-white text-sm font-bold")
                            with ui.element("div").style(
                                    "width:100%;height:6px;background:#1e293b;border-radius:99px"):
                                ui.element("div").style(
                                    f"height:100%;border-radius:99px;background:{color};"
                                    f"width:{round(cnt/mx*100)}%")

                with ui.card().classes("w-full p-4 rounded-xl bg-slate-900 border border-slate-800"):
                    with ui.row().classes("items-center gap-2 mb-3"):
                        ui.label("🏛").classes("text-base leading-none")
                        ui.label("Source Types").classes("text-sm font-bold text-white")
                    with engine.connect() as c:
                        src_rows = c.execute(text(
                            "SELECT COALESCE(source_type,'WEB') st, COUNT(*) cnt "
                            "FROM regulations GROUP BY st"
                        )).fetchall()
                    for st,cnt in src_rows:
                        clr = "#7c3aed" if st=="TELEGRAM_INTEL" else "#3b82f6"
                        with ui.row().classes("w-full items-center justify-between py-1"):
                            ui.label(st).style(f"color:{clr};font-size:11px;font-weight:700")
                            ui.label(str(cnt)).classes("text-slate-300 text-sm font-bold")

    # ══ PAGE: JOBS ════════════════════════════════════════════════

    def _build_jobs(self):
        ui.label("Scrape Job History").classes("text-xl font-black text-white mb-4")
        with ui.card().classes("w-full p-4 rounded-xl bg-slate-900 border border-slate-800"):
            with ui.row().classes("items-center justify-between mb-4 no-wrap"):
                with ui.column().classes("gap-0"):
                    with ui.row().classes("items-center gap-2"):
                        ui.label("⏳")
                        ui.label("Pipeline Sessions").classes("text-base font-bold text-white")
                    ui.label("Per-regulator breakdown: method used, bytes downloaded, docs saved.").classes(
                        "text-xs text-slate-500 mt-0.5")
                ui.button("Refresh",on_click=self._refresh_sessions).props("flat no-caps").classes("text-slate-400 text-xs")
            self._sess_wrap = ui.element("div").classes("w-full")
            self._refresh_sessions()

    # ══ PAGE: SETTINGS ════════════════════════════════════════════

    def _build_settings(self):
        ui.label("Settings").classes("text-xl font-black text-white mb-4")

        with ui.row().classes("w-full gap-5 items-start flex-wrap"):

            # ── Column 1: API + Regulators ────────────────────────
            with ui.column().classes("gap-5").style("min-width:320px;max-width:420px"):

                # API Keys
                with ui.card().classes("w-full p-5 rounded-xl bg-slate-900 border border-slate-800"):
                    with ui.row().classes("items-center gap-2 mb-4"):
                        ui.label("🔑"); ui.label("API Configuration").classes("text-base font-bold text-white")
                    ui.label("OpenAI API Key").classes("text-xs font-bold text-slate-400 mb-1")
                    cur_key = os.getenv("OPENAI_API_KEY","")
                    masked  = ("sk-..." + cur_key[-6:]) if len(cur_key)>10 else ""
                    key_inp = ui.input(placeholder=masked or "sk-...",
                                       password=True, password_toggle_button=True,
                                       ).classes("w-full rr-input").props("outlined dense")
                    ui.label("Leave blank to keep existing key.").classes("text-xs text-slate-600 mt-1 mb-3")
                    ui.button("Save API Key",
                               on_click=lambda: self._save_api_key(
                                   key_inp.value.strip() if key_inp.value.strip() else cur_key
                               )).props("color=primary no-caps unelevated").classes(
                        "w-full rounded-xl font-bold").style("padding:10px")
                    ui.separator().classes("border-slate-700 my-4")
                    with ui.row().classes("items-center gap-2"):
                        ui.element("div").style(
                            "width:7px;height:7px;border-radius:50%;flex-shrink:0;"
                            +("background:#22c55e" if cur_key else "background:#dc2626"))
                        ui.label("API key active — AI extraction enabled" if cur_key else
                                 "No API key — AI disabled").classes(
                            "text-xs "+("text-green-500" if cur_key else "text-red-500"))
                    with ui.row().classes("items-center gap-2 mt-2"):
                        ui.element("div").style(
                            "width:7px;height:7px;border-radius:50%;flex-shrink:0;"
                            +("background:#22c55e" if _HAS_CURL_CFFI else "background:#ca8a04"))
                        ui.label("curl-cffi v0.15 — 403 fallback active"
                                 if _HAS_CURL_CFFI else "curl-cffi missing").classes(
                            "text-xs "+("text-green-500" if _HAS_CURL_CFFI else "text-yellow-500"))

                # Regulators
                with ui.card().classes("w-full p-5 rounded-xl bg-slate-900 border border-slate-800"):
                    with ui.row().classes("items-center gap-2 mb-3"):
                        ui.label("🏦"); ui.label("Active Regulators").classes("text-base font-bold text-white")
                    with SessionLocal() as db:
                        all_regs = db.query(Regulator).order_by(Regulator.jurisdiction).all()
                        reg_data = [(r.id,r.name,r.jurisdiction,r.base_url,r.active) for r in all_regs]
                    self._reg_cbs = {}
                    for rid,name,jur,url,active in reg_data:
                        with ui.card().classes("w-full p-3 mb-2 rounded-lg bg-slate-800/50 border border-slate-700/50"):
                            with ui.row().classes("w-full items-start justify-between no-wrap"):
                                with ui.column().classes("gap-0 flex-1"):
                                    with ui.row().classes("items-center gap-2"):
                                        ui.label(FLAGS.get(jur,"🏳"))
                                        ui.label(name).classes("text-sm font-semibold text-white")
                                    ui.label((url[:52]+"...") if len(url)>52 else url).classes("text-xs text-slate-500 mt-0.5")
                                cb = ui.checkbox(value=active).classes("ml-2")
                                self._reg_cbs[rid] = cb
                    ui.button("Save Regulators",on_click=self._save_regulators).props(
                        "color=primary no-caps unelevated").classes("w-full mt-2 rounded-xl font-bold").style("padding:10px")

            # ── Column 2: Integrations ────────────────────────────
            with ui.column().classes("flex-1 gap-5").style("min-width:320px"):

                with ui.card().classes("w-full p-5 rounded-xl bg-slate-900 border border-slate-800"):
                    with ui.row().classes("items-center gap-2 mb-4"):
                        ui.label("🔔"); ui.label("Integrations & Webhooks").classes("text-base font-bold text-white")
                    ui.label("CRITICAL alerts are auto-dispatched after each pipeline run.").classes(
                        "text-xs text-slate-500 mb-4")

                    wh_inputs: dict = {}

                    # Telegram
                    ui.separator().classes("border-slate-800 mb-3")
                    with ui.row().classes("items-center gap-2 mb-2"):
                        ui.label("✈️"); ui.label("Telegram Bot Alerts").classes("text-sm font-bold text-white")
                    tg_token = ui.input(
                        "Bot Token", value=_wh_get("tg_token"),
                        placeholder="1234567890:AAH..."
                    ).classes("w-full rr-input mb-2").props("outlined dense")
                    tg_chat  = ui.input(
                        "Chat ID / Channel", value=_wh_get("tg_chat_id"),
                        placeholder="-100123456789"
                    ).classes("w-full rr-input mb-2").props("outlined dense")
                    wh_inputs["tg_token"]  = tg_token
                    wh_inputs["tg_chat_id"]= tg_chat
                    ui.button("Send Test Message",
                               on_click=lambda: asyncio.ensure_future(
                                   self._test_telegram(tg_token, tg_chat))
                               ).props("flat no-caps").classes("text-blue-400 text-xs mb-3")

                    # Slack
                    ui.separator().classes("border-slate-800 mb-3")
                    with ui.row().classes("items-center gap-2 mb-2"):
                        ui.label("💬"); ui.label("Slack Webhook").classes("text-sm font-bold text-white")
                    slack_url = ui.input(
                        "Webhook URL", value=_wh_get("slack_url"),
                        placeholder="https://hooks.slack.com/services/..."
                    ).classes("w-full rr-input mb-3").props("outlined dense")
                    wh_inputs["slack_url"] = slack_url

                    # Jira
                    ui.separator().classes("border-slate-800 mb-3")
                    with ui.row().classes("items-center gap-2 mb-2"):
                        ui.label("📌"); ui.label("Jira Auto-Tickets").classes("text-sm font-bold text-white")
                    jira_ep  = ui.input("API Endpoint",value=_wh_get("jira_endpoint"),
                                        placeholder="https://company.atlassian.net").classes(
                        "w-full rr-input mb-2").props("outlined dense")
                    jira_tok = ui.input("API Token (Base64)",value=_wh_get("jira_token"),
                                        password=True,password_toggle_button=True).classes(
                        "w-full rr-input mb-2").props("outlined dense")
                    jira_prj = ui.input("Project Key",value=_wh_get("jira_project"),
                                        placeholder="COMP").classes(
                        "w-full rr-input mb-3").props("outlined dense")
                    wh_inputs["jira_endpoint"] = jira_ep
                    wh_inputs["jira_token"]    = jira_tok
                    wh_inputs["jira_project"]  = jira_prj

                    ui.button("Save All Integration Settings",
                               on_click=lambda: self._save_webhooks(wh_inputs)
                               ).props("color=primary no-caps unelevated").classes(
                        "w-full rounded-xl font-bold mt-2").style("padding:10px")

            # ── Column 3: Company Policies ────────────────────────
            with ui.column().classes("gap-5").style("min-width:300px;max-width:420px"):

                with ui.card().classes("w-full p-5 rounded-xl bg-slate-900 border border-slate-800"):
                    with ui.row().classes("items-center gap-2 mb-2"):
                        ui.label("📄"); ui.label("Internal Policy Conflict Detection").classes("text-base font-bold text-white")
                    ui.label(
                        "Paste internal compliance policies. RegRadar AI will automatically "
                        "compare every new regulation against them and flag conflicts in the table."
                    ).classes("text-xs text-slate-500 mb-4 leading-relaxed")

                    policy_name = ui.input("Policy Name",placeholder="e.g. AML/KYC Policy v3.2").classes(
                        "w-full rr-input mb-2").props("outlined dense")
                    policy_text = ui.textarea("Policy Content (paste text)").classes(
                        "w-full rr-input mb-3").props("outlined dense rows=6")

                    policies_wrap = ui.column().classes("w-full gap-0")

                    ui.button("Upload Policy",
                               on_click=lambda: self._save_policy(policy_name, policy_text, policies_wrap)
                               ).props("color=primary no-caps unelevated").classes(
                        "w-full rounded-xl font-bold mb-4").style("padding:10px")

                    ui.separator().classes("border-slate-700 mb-3")
                    ui.label("Saved Policies").classes("text-xs font-bold text-slate-400 mb-2")
                    self._reload_policies(policies_wrap)


# ══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════

if __name__ in {"__main__","builtins"}:
    app = RegRadarApp()
    app.build_ui()
    ui.run(
        host="0.0.0.0", port=8080,
        title="RegRadar Enterprise",
        favicon="📋",
        reload=False, dark=True,
        storage_secret="rr-enterprise-2026",
    )
