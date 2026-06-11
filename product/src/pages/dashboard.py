"""Dashboard page — KPI cards, trend chart, region chips, regulations grid."""
import json as _json
import threading
import time as _time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from nicegui import APIRouter, app, ui

from config import JURISDICTIONS
from database import reg_repo, job_repo
from src import sidebar

router = APIRouter()

# ── Module-level direct-run state ─────────────────────────────────────────────
# Survives page navigation within the same server process so the UI can restore
# the "running" state when the user comes back from another tab.
_DIRECT: dict = {"running": False, "started": 0.0, "error": ""}

# ── AG-Grid column definitions ────────────────────────────────────────────────
_COL_DEFS = [
    {"field": "jurisdiction", "headerName": "Юр.", "width": 70,
     ":cellRenderer": """(p)=>`<span style='font-size:18px'>${{RU:'🇷🇺',KZ:'🇰🇿',AZ:'🇦🇿',BY:'🇧🇾',UZ:'🇺🇿'}}[p.value]||p.value</span>`"""},
    {"field": "critical_level", "headerName": "Риск", "width": 110,
     ":cellRenderer": """(p)=>{
        const c={CRITICAL:'#dc2626',HIGH:'#ea580c',MEDIUM:'#d97706',LOW:'#059669'}[p.value]||'#6b7280';
        return `<span class='rr-badge' style='background:${c}'>${p.value}</span>`;
     }"""},
    {"field": "effective_date", "headerName": "Дата",     "width": 110},
    {"field": "title",          "headerName": "Название", "flex": 2,
     ":tooltipValueGetter": "p=>p.value"},
    {"field": "summary",        "headerName": "Краткое",  "flex": 3},
    {"field": "confidence",     "headerName": "AI%",      "width": 75,
     ":cellStyle": "p=>({color:p.value>=80?'#4ade80':p.value>=60?'#facc15':'#f87171',fontWeight:600})"},
    {"field": "status", "headerName": "Статус", "width": 110,
     ":cellRenderer": """(p)=>{
        const c={VALIDATED:'#059669',HUMAN_REVIEW:'#d97706',PROCESSING:'#3b82f6'}[p.value]||'#6b7280';
        return `<span class='rr-badge' style='background:${c}'>${p.value}</span>`;
     }"""},
    {"headerName": "", "field": "id", "width": 52, "sortable": False,
     "suppressMenu": True, "suppressMovable": True,
     ":cellRenderer": "()=>`<span style='display:block;text-align:center;color:#1e3a5f;font-size:11px;font-weight:700;cursor:pointer;padding:3px 0' title='History'>⏱</span>`"},
]

_IMPACT_CSS = {
    "CRITICAL": "rr-impact-critical", "HIGH": "rr-impact-high",
    "MEDIUM":   "rr-impact-medium",   "LOW":  "rr-impact-low",
}
_DOT_COLOR = {
    "CRITICAL": "#dc2626", "HIGH": "#f97316",
    "MEDIUM":   "#f59e0b", "LOW":  "#10b981",
}


# ══════════════════════════════════════════════════════════════════════════════
#  DATA REFRESH
# ══════════════════════════════════════════════════════════════════════════════

def _refresh_all(state: dict) -> None:
    jur_filter = state["jurisdiction"] or None
    days       = state["days"]
    level      = state["level"] or None
    refs       = state["refs"]

    rows  = reg_repo.list_recent(days=days, jurisdiction=jur_filter, level=level)
    kpis  = reg_repo.kpis(days=7)
    trend = reg_repo.trend_by_day(days=days)

    # ── KPI labels ────────────────────────────────────────────────
    for key, ref_key in [("total",    "kpi_total"),
                          ("critical", "kpi_crit"),
                          ("review",   "kpi_review"),
                          ("week",     "kpi_week")]:
        if ref_key in refs:
            refs[ref_key].set_text(str(kpis[key]))

    # ── Regulations grid ──────────────────────────────────────────
    if "grid" in refs:
        refs["grid"].options["rowData"] = rows
        refs["grid"].update()

    # ── Trend chart ───────────────────────────────────────────────
    if "chart" in refs:
        series_data: dict = defaultdict(lambda: defaultdict(int))
        for r in trend:
            series_data[r["jurisdiction"]][r["day"]] = r["cnt"]

        days_set = sorted({r["day"] for r in trend})

        # When a region chip is active, isolate that jurisdiction's line so the
        # chart visually reacts; otherwise show all jurisdictions that have data.
        if jur_filter:
            jurs_to_show = [jur_filter] if jur_filter in series_data else []
        else:
            jurs_to_show = [j for j in JURISDICTIONS if j in series_data]

        new_series = [
            {
                "name": jc, "type": "line", "smooth": True,
                "data": [series_data[jc].get(d, 0) for d in days_set],
                "areaStyle": {"opacity": 0.08},
            }
            for jc in jurs_to_show
        ]

        # Mutate in-place — NiceGUI diffs and pushes only changed keys
        refs["chart"].options["xAxis"]["data"] = days_set
        refs["chart"].options["series"] = new_series
        refs["chart"].update()


# ══════════════════════════════════════════════════════════════════════════════
#  MONITORING — SESSION-PERSISTENT POLLING
# ══════════════════════════════════════════════════════════════════════════════

def _start_celery_poll(task_id: str, refs: dict, state: dict) -> None:
    """
    Create a ui.timer that polls AsyncResult every 5 s.
    Runs in NiceGUI context so UI mutations are safe.
    The timer is local to this page render; app.storage.client carries the
    task_id across navigation so a new render can restart the poll.
    """
    timer_box: dict = {"t": None}
    start_ts: float = app.storage.client.get("monitoring_start_ts", _time.time())

    def _done(msg: str, err: bool = False, revoke: bool = False) -> None:
        if timer_box["t"] is not None:
            timer_box["t"].cancel()
            timer_box["t"] = None
        if revoke:
            try:
                from celery_app import celery_app as _ca
                _ca.control.revoke(task_id, terminate=True)
            except Exception:
                pass
        app.storage.client.pop("monitoring_task_id",   None)
        app.storage.client.pop("monitoring_start_ts",  None)
        if "run_btn" in refs:
            refs["run_btn"].enable()
        if "run_status" in refs:
            refs["run_status"].set_text(f"{'❌' if err else '✅'} {msg}")
        if err:
            ui.notify(
                f"Ошибка запуска воркера Celery: {msg}",
                type="negative", position="top-right", timeout=6000,
            )
        try:
            _refresh_all(state)
        except Exception:
            pass

    def _tick() -> None:
        elapsed = int(_time.time() - start_ts)
        try:
            from celery.result import AsyncResult
            result = AsyncResult(task_id)
            s = result.state
        except Exception as exc:
            _done(f"Poll error: {str(exc)[:45]}", err=True)
            return

        if s == "SUCCESS":
            _done(datetime.now().strftime("%H:%M:%S"))
        elif s in ("FAILURE", "REVOKED"):
            err_info = str(result.info)[:55] if result.info else s
            _done(err_info, err=True)
        elif elapsed >= 300:
            _done("Таймаут 5 мин — задача отменена", err=True, revoke=True)
        else:
            if "run_status" in refs:
                refs["run_status"].set_text(f"⏳ Task {task_id[:8]}… ({elapsed}s)")

    timer_box["t"] = ui.timer(5.0, _tick)


def _start_direct_poll(refs: dict, state: dict) -> None:
    """Poll the module-level _DIRECT dict every 3 s (dev / no-Redis fallback)."""
    timer_box: dict = {"t": None}

    def _tick() -> None:
        if not _DIRECT["running"]:
            if timer_box["t"] is not None:
                timer_box["t"].cancel()
                timer_box["t"] = None
            err = _DIRECT.get("error", "")
            if "run_btn" in refs:
                refs["run_btn"].enable()
            if "run_status" in refs:
                if err:
                    refs["run_status"].set_text(f"❌ {err[:55]}")
                    ui.notify(f"Ошибка запуска воркера Celery: {err}", type="negative", timeout=5000)
                else:
                    refs["run_status"].set_text(f"✅ {datetime.now().strftime('%H:%M:%S')}")
            try:
                _refresh_all(state)
            except Exception:
                pass
        else:
            elapsed = int(_time.time() - _DIRECT["started"])
            if elapsed >= 300:
                _DIRECT["running"] = False
                _DIRECT["error"] = "Таймаут 5 мин"
            elif "run_status" in refs:
                refs["run_status"].set_text(f"⏳ Прямой запуск… ({elapsed}s)")

    timer_box["t"] = ui.timer(3.0, _tick)


def _trigger_monitoring(state: dict) -> None:
    from loguru import logger
    refs = state["refs"]

    btn = refs.get("run_btn")
    lbl = refs.get("run_status")

    if btn:
        btn.disable()
    if lbl:
        lbl.set_text("⏳ Запуск…")

    # ── Celery path (production) ──────────────────────────────────
    try:
        from celery_app import scrape_all_regulators_task
        task = scrape_all_regulators_task.apply_async(queue="scraping")
        now = _time.time()
        app.storage.client["monitoring_task_id"]  = task.id
        app.storage.client["monitoring_start_ts"] = now
        if lbl:
            lbl.set_text(f"⏳ Task {task.id[:8]}…")
        _start_celery_poll(task.id, refs, state)
        return

    except Exception as celery_exc:
        # Celery / Redis not reachable — fall through to direct mode
        logger.warning("Celery недоступен ({}): {}", type(celery_exc).__name__, celery_exc)
        ui.notify(
            "Ошибка запуска воркера Celery — переключение на прямой запуск",
            type="warning", position="top-right", timeout=4000,
        )

    # ── Direct fallback (dev / no-Redis) ─────────────────────────
    if _DIRECT["running"]:
        ui.notify("Сбор данных уже запущен", type="info", position="top")
        if btn:
            btn.enable()
        return

    _DIRECT.update({"running": True, "started": _time.time(), "error": ""})
    if lbl:
        lbl.set_text("⏳ Прямой запуск…")

    def _run() -> None:
        try:
            from scheduler_workers import run_all_regulators
            run_all_regulators()
            _DIRECT["error"] = ""
        except Exception as exc:
            logger.error("run_all_regulators: {}", exc)
            _DIRECT["error"] = str(exc)[:80]
        finally:
            _DIRECT["running"] = False

    threading.Thread(target=_run, daemon=True).start()
    _start_direct_poll(refs, state)


# ══════════════════════════════════════════════════════════════════════════════
#  XLSX EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def _export_xlsx(state: dict) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = reg_repo.list_recent(days=state["days"])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Регламенты"
    HDR    = ["#", "Юр.", "Риск", "Дата", "Название", "Краткое", "AI%", "Статус"]
    WIDTHS = [5, 8, 12, 16, 45, 55, 7, 14]
    FILLS  = {"CRITICAL": "DC2626", "HIGH": "EA580C", "MEDIUM": "D97706", "LOW": "059669"}
    for col, (h, w) in enumerate(zip(HDR, WIDTHS), 1):
        cell = ws.cell(1, col, h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    for i, r in enumerate(rows, 2):
        ws.row_dimensions[i].height = 16
        for col, v in enumerate(
            [i - 1, r["jurisdiction"], r["critical_level"], r["effective_date"],
             r["title"], r["summary"], f"{r['confidence']}%", r["status"]], 1
        ):
            cell           = ws.cell(i, col, v)
            cell.alignment = Alignment(vertical="top", wrap_text=(col in {5, 6}))
            cell.font      = Font(size=9)
        risk           = ws.cell(i, 3)
        risk.fill      = PatternFill("solid", fgColor=FILLS.get(r["critical_level"], "6B7280"))
        risk.font      = Font(bold=True, color="FFFFFF", size=9)
        risk.alignment = Alignment(horizontal="center")
    out = Path(f"regrada_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(out)
    ui.download(str(out))


# ══════════════════════════════════════════════════════════════════════════════
#  VERSION HISTORY DIALOG
# ══════════════════════════════════════════════════════════════════════════════

def _build_version_dialog(refs: dict) -> None:
    with ui.dialog().props("persistent=false") as dlg:
        with ui.card().style(
            "background:#080e1c; border:1px solid #1e293b; border-radius:12px;"
            " width:min(760px,93vw); max-height:84vh; padding:0; overflow:hidden;"
            " display:flex; flex-direction:column"
        ):
            with ui.row().classes("items-center justify-between w-full px-5 py-4").style(
                "border-bottom:1px solid #1e293b; flex-shrink:0"
            ):
                with ui.row().classes("items-center gap-2"):
                    ui.icon("history", size="18px").style("color:#60a5fa")
                    dlg_title = ui.label("Version History").classes("text-sm font-bold").style("color:#e2e8f0")
                ui.button(icon="close", on_click=dlg.close).props("flat round dense").style("color:#475569")
            with ui.scroll_area().classes("w-full").style("flex:1; padding:0"):
                dlg_body = ui.column().classes("w-full gap-0 p-4")

    def _open(row_data: dict) -> None:
        if not row_data or "id" not in row_data:
            return
        try:
            from app.infrastructure.db.repository import RegulationRepository
            versions = RegulationRepository().get_versions(row_data["id"])
        except Exception as exc:
            ui.notify(f"Не удалось загрузить историю: {str(exc)[:80]}", type="warning")
            return

        dlg_title.set_text((row_data.get("title") or "Version History")[:70])
        dlg_body.clear()

        with dlg_body:
            if not versions:
                with ui.column().classes("items-center justify-center w-full py-10 gap-2"):
                    ui.icon("history_toggle_off", size="36px").style("color:#1e293b")
                    ui.label("No version history — document unchanged since ingestion.").classes("text-sm").style("color:#334155")
            else:
                for i, ver in enumerate(reversed(versions)):
                    impact  = ver.get("impact_level") or "LOW"
                    is_new  = (i == 0)
                    ver_num = ver.get("version_number", "?")
                    ts      = ver.get("created_at")
                    ts_str  = (ts.strftime("%Y-%m-%d  %H:%M UTC") if hasattr(ts, "strftime") else str(ts or "")[:16])
                    fp      = ver.get("summary_hash") or ""

                    with ui.element("div").classes("rr-timeline-item"):
                        ui.element("div").classes("rr-timeline-dot").style(
                            f"background:{_DOT_COLOR.get(impact,'#334155')}"
                            + ("; box-shadow:0 0 0 3px rgba(59,130,246,.2)" if is_new else "")
                        )
                        with ui.row().classes("items-center gap-2 mb-2 flex-wrap"):
                            ui.html(f'<span style="background:#1e293b;color:#94a3b8;font-family:monospace;font-size:11px;font-weight:700;padding:2px 8px;border-radius:4px">v{ver_num}</span>')
                            ui.label(ts_str).classes("text-xs").style("color:#475569")
                            if fp:
                                ui.html(f'<span style="color:#94a3b8;font-family:monospace;font-size:10px" title="SHA-256">#{fp}</span>')
                            if is_new:
                                ui.badge("LATEST", color="blue-grey").classes("text-xs")

                        if ver.get("delta_json"):
                            try:
                                delta   = _json.loads(ver["delta_json"])
                                imp     = delta.get("impact_level", impact)
                                css_cls = _IMPACT_CSS.get(imp, "rr-impact-low")
                                with ui.row().classes("items-center gap-2 mb-3"):
                                    ui.html(f'<span class="rr-impact-pill {css_cls}">● {imp} IMPACT</span>')
                                changes = delta.get("critical_changes", [])
                                if changes:
                                    ui.label("MATERIAL CHANGES").classes("text-xs font-bold mb-2").style("color:#6366f1;letter-spacing:.1em")
                                    for chg in changes[:6]:
                                        with ui.row().classes("items-start gap-2 mb-1"):
                                            ui.html('<span style="color:#3b82f6;flex-shrink:0;line-height:1.5;font-size:13px">•</span>')
                                            ui.label(chg[:160]).classes("text-xs").style("color:#94a3b8;line-height:1.55")
                                action = delta.get("business_action_required", "")
                                if action:
                                    with ui.element("div").style(
                                        "background:rgba(14,28,56,.8);border-left:3px solid #1d4ed8;"
                                        "border-radius:0 6px 6px 0;padding:10px 14px;margin-top:10px"
                                    ):
                                        ui.label("COMPLIANCE ACTION").classes("text-xs font-bold mb-1").style("color:#6366f1;letter-spacing:.1em")
                                        ui.label(action).classes("text-xs").style("color:#cbd5e1;font-style:italic;line-height:1.65")
                            except Exception:
                                ui.label("Delta analysis pending…").classes("text-xs").style("color:#1e293b;font-style:italic")
                        elif ver_num == 1:
                            ui.label("Initial ingestion — baseline version established.").classes("text-xs").style("color:#1e293b;font-style:italic")
                        else:
                            ui.label("Delta analysis queued — check back shortly.").classes("text-xs").style("color:#1e293b;font-style:italic")

        dlg.open()

    refs["_open_version"] = _open


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE
# ══════════════════════════════════════════════════════════════════════════════

@router.page("/")
def dashboard_page() -> None:
    dark_mode = ui.dark_mode(value=True)
    state: dict = {"days": 30, "jurisdiction": "", "level": "", "refs": {}}
    refs = state["refs"]

    sidebar.render("dashboard", dark_mode)

    with ui.column().classes("w-full p-0"):

        # ── Header bar ────────────────────────────────────────────
        with ui.row().classes("w-full items-center justify-between px-6 py-3").style(
            "background:#1e293b; border-bottom:1px solid #334155"
        ):
            with ui.row().classes("items-center gap-2"):
                ui.html('<span style="font-size:13px;color:#94a3b8">Regulatory Signal Intelligence</span>')
                ui.badge("BETA", color="blue-grey").classes("text-xs")
            with ui.row().classes("items-center gap-3"):
                ui.button(icon="refresh", on_click=lambda: _refresh_all(state)).props("flat round").style("color:#94a3b8")
                refs["run_btn"] = ui.button(
                    "▶ Запустить мониторинг",
                    on_click=lambda: _trigger_monitoring(state),
                ).props("color=primary no-caps").classes("text-sm")
                refs["run_status"] = ui.label("").classes("text-xs").style(
                    "color:#94a3b8; min-width:200px"
                )

        with ui.column().classes("w-full p-6"):

            # ── KPI row ───────────────────────────────────────────
            with ui.row().classes("gap-4 w-full flex-wrap mb-6"):
                for bg, icon, label, ref_key in [
                    ("linear-gradient(135deg,#1e3a8a,#1e40af)", "article",       "Всего документов", "kpi_total"),
                    ("linear-gradient(135deg,#7f1d1d,#991b1b)", "warning_amber", "Критических (7д)", "kpi_crit"),
                    ("linear-gradient(135deg,#78350f,#92400e)", "person_search", "На ревью",         "kpi_review"),
                    ("linear-gradient(135deg,#064e3b,#065f46)", "trending_up",   "Новых за 7 дней",  "kpi_week"),
                ]:
                    with ui.card().classes("rr-kpi shadow-lg flex-1").style(f"background:{bg}"):
                        ui.icon(icon, size="28px").style("color:rgba(255,255,255,.6)")
                        ui.label(label).classes("text-xs font-semibold uppercase tracking-widest mt-1").style("color:rgba(255,255,255,.6)")
                        refs[ref_key] = ui.label("—").classes("text-4xl font-bold mt-1 text-white")

            # ── Trend chart ───────────────────────────────────────
            with ui.card().classes("w-full shadow-md mb-5 p-4").style("background:#1e293b"):
                with ui.row().classes("items-center gap-2 mb-2"):
                    ui.icon("show_chart", size="20px").style("color:#60a5fa")
                    ui.label("Динамика обнаружений").classes("text-sm font-semibold text-slate-300")
                refs["chart"] = ui.echart({
                    "tooltip":  {"trigger": "axis"},
                    "legend":   {"data": list(JURISDICTIONS.keys()), "textStyle": {"color": "#94a3b8"}},
                    "grid":     {"left": "3%", "right": "4%", "bottom": "3%", "containLabel": True},
                    "xAxis":    {"type": "category", "data": [], "axisLabel": {"color": "#94a3b8"}},
                    "yAxis":    {"type": "value", "axisLabel": {"color": "#94a3b8"},
                                 "splitLine": {"lineStyle": {"color": "#1e293b"}}},
                    "series":   [],
                    "backgroundColor": "transparent",
                }).classes("w-full").style("height:220px")

            # ── Region selector ───────────────────────────────────
            with ui.card().classes("rr-card w-full mb-3 px-4 py-3").style("background:#0d1424"):
                with ui.row().classes("items-center gap-3 flex-wrap"):
                    ui.icon("public", size="16px").style("color:#475569")
                    ui.html('<span style="font-size:11px;font-weight:700;letter-spacing:.08em;color:#475569">REGION</span>')
                    ui.element("div").style("width:1px;height:18px;background:#1e293b;flex-shrink:0")

                    _REGIONS = [
                        ("",   "🌐  Global"), ("RU", "🇷🇺  RU"), ("KZ", "🇰🇿  KZ"),
                        ("AZ", "🇦🇿  AZ"),   ("BY", "🇧🇾  BY"), ("UZ", "🇺🇿  UZ"),
                    ]
                    region_btns: dict = {}

                    def _select_region(code: str) -> None:
                        state["jurisdiction"] = code
                        for k, b in refs["region_btns"].items():
                            if k == code:
                                b.classes(add="rr-region-active", remove="rr-region-inactive")
                            else:
                                b.classes(add="rr-region-inactive", remove="rr-region-active")
                        _refresh_all(state)

                    for _code, _lbl in _REGIONS:
                        _btn = (
                            ui.button(_lbl, on_click=lambda c=_code: _select_region(c))
                            .classes("rr-region-btn rr-region-inactive")
                            .props("flat no-caps no-shadow")
                        )
                        if _code == "":
                            _btn.classes(add="rr-region-active", remove="rr-region-inactive")
                        region_btns[_code] = _btn
                    refs["region_btns"] = region_btns

            # ── Secondary filters ─────────────────────────────────
            with ui.row().classes("items-center gap-3 mb-3 flex-wrap"):
                ui.icon("filter_list", size="18px").style("color:#475569")
                level_sel = ui.select(
                    options={"": "All Risk Levels", "CRITICAL": "🔴 Critical",
                             "HIGH": "🟠 High", "MEDIUM": "🟡 Medium", "LOW": "🟢 Low"},
                    value="", label="Risk Level",
                ).classes("w-40").props("outlined dense dark")
                level_sel.on("update:model-value",
                    lambda e: (state.update({"level": e.args}), _refresh_all(state)))

                days_sel = ui.select(
                    options={7: "7 days", 30: "30 days", 90: "90 days", 365: "1 year"},
                    value=30, label="Period",
                ).classes("w-32").props("outlined dense dark")
                days_sel.on("update:model-value",
                    lambda e: (state.update({"days": int(e.args)}), _refresh_all(state)))

                ui.button(icon="file_download", on_click=lambda: _export_xlsx(state)).props(
                    "outline no-caps color=teal"
                ).classes("ml-auto").tooltip("Export Excel")

            # ── Regulations grid ──────────────────────────────────
            refs["grid"] = ui.aggrid({
                "columnDefs": _COL_DEFS,
                "rowData": [],
                "rowClassRules": {
                    "border-l-4 border-red-600":    "params.data.critical_level==='CRITICAL'",
                    "border-l-4 border-orange-500": "params.data.critical_level==='HIGH'",
                },
                "pagination": True, "paginationPageSize": 15,
                "domLayout": "autoHeight", "theme": "quartz",
                "enableCellTextSelection": True, "tooltipShowDelay": 300,
            }).classes("w-full")

    # ── Version history dialog ────────────────────────────────────
    _build_version_dialog(refs)
    refs["grid"].on("rowClicked", lambda e: refs["_open_version"](e.args.get("data") or {}))

    # ── Restore monitoring state after navigation ─────────────────
    # app.storage.client persists the task id for the browser tab session.
    # _DIRECT persists direct-run state at module level within the process.
    stored_task_id = app.storage.client.get("monitoring_task_id", "")
    if stored_task_id:
        refs["run_btn"].disable()
        elapsed = int(_time.time() - app.storage.client.get("monitoring_start_ts", _time.time()))
        refs["run_status"].set_text(f"⏳ Task {stored_task_id[:8]}… ({elapsed}s, восст.)")
        _start_celery_poll(stored_task_id, refs, state)
    elif _DIRECT["running"]:
        refs["run_btn"].disable()
        elapsed = int(_time.time() - _DIRECT["started"])
        refs["run_status"].set_text(f"⏳ Прямой запуск… ({elapsed}s, восст.)")
        _start_direct_poll(refs, state)

    _refresh_all(state)
